"""Preflight validator tests.

One case per rejection reason, plus the positive control. The positive
control matters as much as the rejections: a validator that rejects
everything would pass every rejection test.
"""

import hashlib
import shutil

import pytest

from materia.preflight import PreflightRejected, Reason, preflight

# fixture name -> the code the user should see
REJECTIONS = [
    ("vba", "VBA_PRESENT"),
    ("external_link_part", "EXTERNAL_LINK"),
    ("external_link_formula", "EXTERNAL_LINK"),
    ("array_formula", "ARRAY_FORMULA"),
    ("circular", "CIRCULAR_REFERENCE"),
    ("circular_via_range", "CIRCULAR_REFERENCE"),
    ("circular_cross_sheet", "CIRCULAR_REFERENCE"),
    ("unsupported_function", "UNSUPPORTED_FUNCTION(VLOOKUP)"),
    ("unsupported_function_lookalike", "UNSUPPORTED_FUNCTION(LOG10)"),
    ("unsupported_function_nested", "UNSUPPORTED_FUNCTION(SQRT)"),
    ("defined_name", "DEFINED_NAME"),
    ("unparseable_formula", "UNPARSEABLE_FORMULA"),
]


@pytest.mark.parametrize("name,expected_code", REJECTIONS)
def test_rejects_with_the_right_reason(workbooks, name, expected_code):
    with pytest.raises(PreflightRejected) as raised:
        preflight(workbooks[name])
    assert raised.value.code == expected_code


@pytest.mark.parametrize("name,_code", REJECTIONS)
def test_rejection_message_is_actionable(workbooks, name, _code):
    """Every rejection names the reason, and formula level ones name a cell."""
    with pytest.raises(PreflightRejected) as raised:
        preflight(workbooks[name])
    rejection = raised.value
    assert rejection.code in rejection.message
    assert len(rejection.detail) > 20
    if rejection.reason in (Reason.ARRAY_FORMULA, Reason.UNSUPPORTED_FUNCTION):
        assert "!" in rejection.location


def test_clean_workbook_is_accepted(workbooks):
    report = preflight(workbooks["clean"])
    assert report.sheet_names == ["Assumptions", "Model"]
    assert report.formula_count == 16
    assert report.value_cell_count == 8


def test_clean_workbook_uses_every_supported_function(workbooks):
    """The positive control has to actually exercise the grammar.

    Without this, the clean fixture could quietly drift into something
    trivial and still pass, and then it would prove nothing.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(workbooks["clean"], read_only=True)
    formulas = " ".join(
        cell.value
        for sheet in workbook.sheetnames
        for row in workbook[sheet].iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    workbook.close()
    for function in ("SUM", "AVERAGE", "MIN", "MAX", "IF", "ROUND", "ABS", "SUMIF"):
        assert f"{function}(" in formulas, f"{function} missing from the control"


def test_a_defined_name_is_rejected_by_name(workbooks):
    """The message has to say which name, or the user cannot go and find it."""
    with pytest.raises(PreflightRejected) as raised:
        preflight(workbooks["defined_name"])
    assert raised.value.reason is Reason.DEFINED_NAME
    assert "Q1" in raised.value.message


def test_a_print_area_alone_is_not_a_defined_name(workbooks):
    """Built in _xlnm. names never appear in a formula.

    Rejecting a workbook for having a print area would be over-rejection, and
    print areas are on almost every real model.
    """
    report = preflight(workbooks["print_area_only"])
    assert report.formula_count == 1


def test_deep_chains_are_not_mistaken_for_cycles(workbooks):
    """The negative control for circular detection.

    Real forecast models are deep by construction. Flagging depth as a loop
    would reject every workbook the tool exists to audit.
    """
    report = preflight(workbooks["deep_chain"])
    assert report.formula_count == 15


def test_unsupported_function_names_the_function(workbooks):
    with pytest.raises(PreflightRejected) as raised:
        preflight(workbooks["unsupported_function"])
    rejection = raised.value
    assert rejection.reason is Reason.UNSUPPORTED_FUNCTION
    assert rejection.function == "VLOOKUP"
    assert "VLOOKUP" in rejection.message


def test_circular_message_shows_the_loop(workbooks):
    with pytest.raises(PreflightRejected) as raised:
        preflight(workbooks["circular"])
    assert "Sheet!A1" in raised.value.message
    assert "Sheet!A2" in raised.value.message


def test_input_workbook_is_never_modified(workbooks, tmp_path):
    """Data flow invariant 1 in docs/ARCHITECTURE.md: opened read only."""
    original = tmp_path / "subject.xlsx"
    shutil.copy(workbooks["clean"], original)
    before = hashlib.sha256(original.read_bytes()).hexdigest()
    preflight(original)
    assert hashlib.sha256(original.read_bytes()).hexdigest() == before


def test_workbook_with_no_formulas_is_accepted(tmp_path):
    """A workbook of pure constants has nothing to reject."""
    import openpyxl

    path = tmp_path / "constants.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = 1
    workbook.active["A2"] = "a label"
    workbook.save(path)

    report = preflight(path)
    assert report.formula_count == 0
    assert report.value_cell_count == 2


def test_a_file_that_is_not_a_workbook_is_not_a_preflight_rejection(tmp_path):
    """PreflightRejected means a real workbook we cannot evaluate.

    A .csv renamed to .xlsx is a different problem, and reporting it under one
    of the five reason codes would tell the user something untrue.
    """
    path = tmp_path / "not_really.xlsx"
    path.write_text("a,b,c\n1,2,3\n")

    with pytest.raises(ValueError) as raised:
        preflight(path)
    assert not isinstance(raised.value, PreflightRejected)


def test_missing_file_raises_file_not_found(tmp_path):
    with pytest.raises(FileNotFoundError):
        preflight(tmp_path / "nope.xlsx")


class TestParserIntegration:
    """Preflight runs the real parser, so acceptance means the pipeline can
    read the workbook. These check that a parse failure maps to the right
    reason rather than to a single catch all."""

    def test_an_unsupported_function_is_named(self, workbooks):
        with pytest.raises(PreflightRejected) as raised:
            preflight(workbooks["unsupported_function"])
        assert raised.value.code == "UNSUPPORTED_FUNCTION(VLOOKUP)"

    def test_a_malformed_formula_is_not_reported_as_an_unsupported_function(
        self, workbooks
    ):
        """=A1+ has no function in it. Reporting one would be misleading."""
        with pytest.raises(PreflightRejected) as raised:
            preflight(workbooks["unparseable_formula"])
        assert raised.value.reason is Reason.UNPARSEABLE_FORMULA
        assert raised.value.function is None
        assert "A2" in raised.value.location

    def test_every_formula_in_an_accepted_workbook_parses(self, workbooks):
        """The invariant the parse step exists to provide.

        Previously this held by test on a few fixtures. Now preflight parses
        every formula itself, so acceptance guarantees it.
        """
        from materia.formula import parse_formula
        from materia.parse import read_formulas

        for name in ("clean", "copied_formulas", "deep_chain", "print_area_only"):
            preflight(workbooks[name])
            for cell in read_formulas(workbooks[name]):
                parse_formula(cell.formula)
