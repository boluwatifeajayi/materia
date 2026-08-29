"""Corpus generator tests.

The important one is test_the_engine_reproduces_every_cached_value. The
workbook is built twice from the same design, once as Excel formulas and once
as a plain Python month loop, and the second result is written into the file
as the cached values. That test asserts the recompute engine, reading the
formulas, lands on the numbers the loop produced.

A generated workbook has no Excel written values in it, so checking the engine
against a file it produced itself would prove nothing. Two independent
implementations agreeing is a real signal, and either one being wrong shows up
as a failure here.
"""

import hashlib

import openpyxl
import pytest

from materia.corpus import generate
from materia.corpus.layout import (
    DECLARED_OUTPUTS,
    MONTHS,
    PL_SHEET,
    TOTAL,
    month_column,
)
from materia.formula import parse_formula
from materia.parse import read_formulas
from materia.preflight import preflight
from materia.recompute import Model

SEED = 20260828


@pytest.fixture(scope="module")
def workbook(tmp_path_factory):
    return generate(tmp_path_factory.mktemp("corpus") / "C01.xlsx", SEED)[0]


@pytest.fixture(scope="module")
def cached(workbook) -> dict[str, float]:
    """Every numeric value openpyxl reads back out of the file."""
    values = {}
    book = openpyxl.load_workbook(workbook, data_only=True)
    for name in book.sheetnames:
        for row in book[name].iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    values[f"{name}!{cell.coordinate}"] = float(cell.value)
    book.close()
    return values


class TestTheWorkbookIsUsable:
    def test_it_passes_preflight(self, workbook):
        report = preflight(workbook)
        assert report.sheet_names == [
            "Assumptions",
            "Revenue",
            "Costs",
            PL_SHEET,
            "Valuation",
        ]

    def test_every_formula_parses(self, workbook):
        formulas = read_formulas(workbook)
        assert formulas
        for cell in formulas:
            parse_formula(cell.formula)

    def test_the_formula_count_is_in_range(self, workbook):
        """docs/EVALUATION.md section 2 asks for 400 to 1500 formulas."""
        count = preflight(workbook).formula_count
        assert 400 <= count <= 1500, count

    def test_it_has_twenty_four_monthly_columns(self, workbook):
        """Months run C to Z, then the column after them is the totals
        column, so the month row has to stop exactly where the totals
        column starts."""
        book = openpyxl.load_workbook(workbook)
        revenue, profit_and_loss = book["Revenue"], book[PL_SHEET]

        assert revenue[f"{month_column(1)}3"].value is not None
        assert revenue[f"{month_column(MONTHS)}3"].value is not None
        assert revenue[f"{month_column(MONTHS + 1)}3"].value is None

        assert month_column(MONTHS + 1) == TOTAL
        assert profit_and_loss[f"{TOTAL}3"].value == "Total"
        book.close()

    def test_it_uses_every_function_in_the_grammar(self, workbook):
        """A corpus that only used arithmetic would leave most of the engine
        untested by the thing it exists to test."""
        text = " ".join(cell.formula for cell in read_formulas(workbook))
        for function in ("SUM", "AVERAGE", "MIN", "MAX", "IF", "ROUND", "ABS", "SUMIF"):
            assert f"{function}(" in text, f"{function} is not exercised by the corpus"


class TestDeterminism:
    def test_the_same_seed_gives_a_byte_identical_file(self, tmp_path):
        first = generate(tmp_path / "first.xlsx", SEED)[0]
        second = generate(tmp_path / "second.xlsx", SEED)[0]
        assert (
            hashlib.sha256(first.read_bytes()).hexdigest()
            == hashlib.sha256(second.read_bytes()).hexdigest()
        )

    def test_a_different_seed_gives_a_different_file(self, tmp_path):
        first = generate(tmp_path / "first.xlsx", SEED)[0]
        other = generate(tmp_path / "other.xlsx", SEED + 1)[0]
        assert first.read_bytes() != other.read_bytes()

    def test_regenerating_over_an_existing_file_is_stable(self, tmp_path):
        path = tmp_path / "same.xlsx"
        generate(path, SEED)
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        generate(path, SEED)
        assert hashlib.sha256(path.read_bytes()).hexdigest() == digest


class TestTheEngineAgreesWithTheModel:
    def test_the_file_actually_carries_cached_values(self, cached):
        """Without this the comparison below could pass by comparing nothing.

        openpyxl cannot write a formula and its result together, so the
        generator injects the values into the saved XML. If that ever stops
        working, this fails rather than the comparison silently going green.
        """
        assert len(cached) > 700
        assert f"{PL_SHEET}!{TOTAL}15" in cached

    def test_every_formula_cell_carries_a_value(self, workbook, cached):
        """No formula cell may be left without one.

        A missing value does not fail loudly on its own: the comparison below
        would skip that cell and still pass. So the generator raises, and this
        checks the two counts line up.
        """
        formulas = {cell.address for cell in read_formulas(workbook)}
        assert formulas
        assert formulas <= set(cached)

    def test_a_formula_with_no_computed_value_is_refused(self, tmp_path):
        """The guard itself, exercised by withholding one value."""
        from materia.corpus.generate import (
            Assumptions,
            MissingComputedValue,
            build_workbook,
            compute_values,
            save,
        )

        assumptions = Assumptions.from_seed(SEED)
        values = compute_values(assumptions)
        values.cells.pop(f"{PL_SHEET}!{TOTAL}15")
        with pytest.raises(MissingComputedValue, match="AA15"):
            save(build_workbook(assumptions), tmp_path / "broken.xlsx", values)

    def test_the_engine_reproduces_every_cached_value(self, workbook, cached):
        """The cross check the whole task exists for."""
        model = Model.load(workbook)
        compared = 0
        mismatches = []
        for address, expected in cached.items():
            actual = model.value(address)
            if not isinstance(actual, (int, float)):
                continue
            compared += 1
            if abs(float(actual) - expected) > 1e-6:
                mismatches.append((address, expected, float(actual)))

        assert compared > 700, f"only {compared} cells compared"
        assert not mismatches, mismatches[:10]

    def test_the_declared_outputs_are_present_and_numeric(self, workbook, cached):
        model = Model.load(workbook, outputs=DECLARED_OUTPUTS)
        for output in DECLARED_OUTPUTS:
            assert output in cached
            assert isinstance(model.value(output), (int, float))

    def test_a_patch_moves_the_declared_outputs(self, workbook):
        """A corpus workbook has to be one where errors have consequences.

        If the outputs did not move, every mutation would be immaterial and
        the corpus could not test the materiality gate at all.
        """
        model = Model.load(workbook, outputs=DECLARED_OUTPUTS)
        result = model.patch("Assumptions!B4", 1)  # opening customers
        for output in DECLARED_OUTPUTS:
            assert result.outputs[output].delta != 0.0


class TestTheModelIsRealistic:
    def test_it_is_profitable_and_growing(self, cached):
        """A forecast nobody would apply a multiple to is not a realistic
        subject for an audit."""
        first = cached[f"{PL_SHEET}!{month_column(1)}15"]
        last = cached[f"{PL_SHEET}!{month_column(MONTHS)}15"]
        assert first > 0
        assert last > first

    def test_revenue_grows_every_month(self, cached):
        revenue = [cached[f"{PL_SHEET}!{month_column(m)}5"] for m in range(1, MONTHS + 1)]
        assert all(later > earlier for earlier, later in zip(revenue, revenue[1:]))

    def test_the_total_row_equals_the_sum_of_the_months(self, cached):
        months = sum(cached[f"{PL_SHEET}!{month_column(m)}15"] for m in range(1, MONTHS + 1))
        assert cached[f"{PL_SHEET}!{TOTAL}15"] == pytest.approx(months)

    @pytest.mark.parametrize("seed", [1, 20260828, 99999, 424242])
    def test_any_seed_produces_a_sound_workbook(self, tmp_path, seed):
        """The generator has to hold up across seeds, since T08 needs twelve."""
        path = generate(tmp_path / f"seed_{seed}.xlsx", seed)[0]
        report = preflight(path)
        assert 400 <= report.formula_count <= 1500
        model = Model.load(path, outputs=DECLARED_OUTPUTS)
        for output in DECLARED_OUTPUTS:
            assert isinstance(model.value(output), (int, float))


class TestTheTwelveWorkbookCorpus:
    """docs/EVALUATION.md section 2: eight seeded, two clean controls, two
    hard cases."""

    def test_all_twelve_are_generated(self, corpus):
        directory, manifest = corpus
        assert [entry["id"] for entry in manifest["workbooks"]] == [
            f"C{index:02d}" for index in range(1, 13)
        ]
        for entry in manifest["workbooks"]:
            assert (directory / entry["file"]).exists()

    def test_the_roles_match_the_corpus_table(self, corpus):
        _, manifest = corpus
        roles = {entry["id"]: entry["role"] for entry in manifest["workbooks"]}
        for identifier in [f"C{index:02d}" for index in range(1, 9)]:
            assert roles[identifier] == "seeded"
        assert roles["C09"] == roles["C10"] == "clean_control"
        assert roles["C11"] == roles["C12"] == "hard_case"

    def test_every_workbook_passes_preflight(self, corpus):
        directory, manifest = corpus
        for entry in manifest["workbooks"]:
            report = preflight(directory / entry["file"])
            assert report.formula_count == entry["formula_count"]

    def test_every_seed_is_different(self, corpus):
        _, manifest = corpus
        seeds = [entry["seed"] for entry in manifest["workbooks"]]
        assert len(set(seeds)) == len(seeds)

    def test_every_workbook_is_a_viable_business(self, corpus):
        """A negative enterprise value makes a percentage change in it close
        to meaningless, so the materiality gate could not be tested on one."""
        _, manifest = corpus
        for entry in manifest["workbooks"]:
            for output, value in entry["output_values"].items():
                assert value > 0, f"{entry['id']} {output} is {value}"

    def test_the_engine_agrees_with_every_workbook(self, corpus):
        """The T07 cross check, across the whole corpus rather than one file."""
        directory, manifest = corpus
        for entry in manifest["workbooks"]:
            path = directory / entry["file"]
            book = openpyxl.load_workbook(path, data_only=True)
            cached = {
                f"{name}!{cell.coordinate}": float(cell.value)
                for name in book.sheetnames
                for row in book[name].iter_rows()
                for cell in row
                if isinstance(cell.value, (int, float))
            }
            book.close()

            model = Model.load(path)
            for address, expected in cached.items():
                actual = model.value(address)
                if isinstance(actual, (int, float)):
                    assert abs(float(actual) - expected) < 1e-6, f"{entry['id']} {address}"


class TestChecksums:
    def test_checksums_are_stable_across_two_runs(self, tmp_path):
        from materia.corpus.build import CHECKSUMS_NAME, build_corpus

        first = tmp_path / "first"
        second = tmp_path / "second"
        build_corpus(first)
        build_corpus(second)
        assert (first / CHECKSUMS_NAME).read_text() == (second / CHECKSUMS_NAME).read_text()

    def test_check_passes_on_a_fresh_build(self, corpus):
        from materia.corpus.build import check_corpus

        result = check_corpus(corpus[0])
        assert result.ok
        assert len(result.matched) == 12

    def test_check_notices_an_edited_workbook(self, corpus_copy):
        """The point of committing checksums. A workbook that drifted has to
        be caught, or a judge is scoring a different corpus from ours."""
        from materia.corpus.build import check_corpus

        (corpus_copy / "C03.xlsx").write_bytes(b"not the workbook you committed")
        result = check_corpus(corpus_copy)
        assert not result.ok
        assert result.mismatched == ["C03.xlsx"]

    def test_check_notices_a_missing_workbook(self, corpus_copy):
        from materia.corpus.build import check_corpus

        (corpus_copy / "C07.xlsx").unlink()
        result = check_corpus(corpus_copy)
        assert not result.ok
        assert result.missing == ["C07.xlsx"]

    def test_blank_lines_in_the_checksums_file_are_tolerated(self, corpus_copy):
        """A judge may open this file. A stray blank line should not crash the
        check with a split error."""
        from materia.corpus.build import CHECKSUMS_NAME, check_corpus

        path = corpus_copy / CHECKSUMS_NAME
        path.write_text(path.read_text() + "\n\n")
        assert check_corpus(corpus_copy).ok

    def test_check_without_a_build_says_so(self, tmp_path):
        from materia.corpus.build import check_corpus

        with pytest.raises(FileNotFoundError):
            check_corpus(tmp_path)


class TestC10LegitimatePatternBreaks:
    """C10 is the workbook that breaks naive tools. Each break is correct
    model building that a structural detector will flag anyway."""

    @staticmethod
    @pytest.fixture
    def c10(corpus):
        directory, manifest = corpus
        entry = next(e for e in manifest["workbooks"] if e["id"] == "C10")
        return directory / entry["file"], entry

    def test_it_records_exactly_three_breaks(self, c10):
        _, entry = c10
        kinds = {item["kind"] for item in entry["legitimate_breaks"]}
        assert kinds == {"hardcoded_actuals", "first_period", "manual_override"}

    def test_every_break_explains_itself(self, c10):
        """The explanation is what the agent has to find. A break with no
        readable reason is indistinguishable from an error."""
        _, entry = c10
        for item in entry["legitimate_breaks"]:
            assert len(item["why"]) > 40

    def test_the_actuals_row_is_hardcoded(self, c10):
        path, entry = c10
        book = openpyxl.load_workbook(path)
        actuals = next(
            item for item in entry["legitimate_breaks"] if item["kind"] == "hardcoded_actuals"
        )
        for address in actuals["cells"]:
            sheet, coordinate = address.split("!")
            assert isinstance(book[sheet][coordinate].value, (int, float))
        # and the forecast months next to it are still formulas
        assert str(book["Revenue"]["F15"].value).startswith("=")
        book.close()

    def test_the_actuals_are_labelled_as_actuals(self, c10):
        """Evidence the agent can read. Without it the hardcode looks like a
        paste over a formula, which is exactly mutation family M1."""
        path, _ = c10
        book = openpyxl.load_workbook(path)
        assert book["Revenue"]["C2"].value == "Actual"
        assert book["Revenue"]["F2"].value == "Forecast"
        book.close()

    def test_the_manual_override_carries_a_comment(self, c10):
        path, entry = c10
        book = openpyxl.load_workbook(path)
        override = next(
            item for item in entry["legitimate_breaks"] if item["kind"] == "manual_override"
        )
        sheet, coordinate = override["cells"][0].split("!")
        cell = book[sheet][coordinate]
        assert isinstance(cell.value, (int, float))
        assert cell.comment is not None
        assert "board" in cell.comment.text.lower()
        book.close()

    def test_it_is_still_a_clean_control(self, c10):
        path, entry = c10
        assert entry["mutations"] == []
        preflight(path)

    def test_it_is_deterministic_with_comments_in_it(self, tmp_path):
        """Comments add parts to the zip, which is where determinism usually
        breaks."""
        first = generate(tmp_path / "a.xlsx", 1234, legitimate_breaks=True)[0]
        second = generate(tmp_path / "b.xlsx", 1234, legitimate_breaks=True)[0]
        assert first.read_bytes() == second.read_bytes()


class TestManifest:
    @staticmethod
    @pytest.fixture
    def manifest(corpus):
        return corpus[1]

    def test_it_validates(self, manifest):
        from materia.corpus.manifest import validate_manifest

        validate_manifest(manifest)

    def test_it_round_trips_through_disk(self, manifest, tmp_path):
        from materia.corpus.manifest import read_manifest, write_manifest

        path = write_manifest(tmp_path / "manifest.json", manifest)
        assert read_manifest(path) == manifest

    def test_it_is_written_sorted_so_a_diff_is_readable(self, corpus):
        import json

        from materia.corpus.build import MANIFEST_NAME

        text = (corpus[0] / MANIFEST_NAME).read_text()
        assert text == json.dumps(json.loads(text), indent=2, sort_keys=True) + "\n"

    @pytest.mark.parametrize(
        "mutate,expected",
        [
            (lambda m: m.pop("seed"), "missing 'seed'"),
            (lambda m: m.update(version=99), "unsupported version"),
            (lambda m: m.update(declared_outputs=["P&L!AA15"]), "two Sheet!Cell"),
            (lambda m: m["workbooks"][0].update(role="nonsense"), "unknown role"),
            (lambda m: m["workbooks"][0].update(sha256="abc"), "not a sha256"),
            (lambda m: m["workbooks"][0].update(id="nope"), "bad workbook id"),
            (lambda m: m["workbooks"][0].pop("mutations"), "missing 'mutations'"),
            (lambda m: m["workbooks"][0]["output_values"].popitem(), "exactly the declared"),
            (lambda m: m["workbooks"][0].update(formula_count=0), "must be positive"),
            (lambda m: m["workbooks"][0].update(file="wrong.xlsx"), "file name does not match"),
        ],
    )
    def test_it_rejects_a_broken_manifest(self, manifest, mutate, expected):
        """A wrong manifest makes every metric in the submission wrong, so it
        is checked rather than trusted."""
        import copy

        from materia.corpus.manifest import InvalidManifest, validate_manifest

        broken = copy.deepcopy(manifest)
        mutate(broken)
        with pytest.raises(InvalidManifest, match=expected):
            validate_manifest(broken)
