"""Fixture workbook builders.

Generated rather than committed, so every fixture is inspectable as code and
nothing in the repo is an opaque binary. Run this module directly to write
them all somewhere you can open them:

    python tests/fixtures/build.py /tmp/materia-fixtures

Two of the rejection cases cannot be authored by openpyxl: it has no way to
write a VBA project or an external link part. Those two are built by writing a
normal workbook and then injecting the relevant part into the .xlsx zip, which
is exactly the structure preflight looks for.
"""

from __future__ import annotations

import shutil
import sys
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula


def _inject_zip_part(path: Path, part_name: str, content: bytes) -> None:
    """Add one part to an existing .xlsx, rewriting the zip in place."""
    temporary = path.with_suffix(".tmp.zip")
    with zipfile.ZipFile(path) as source, zipfile.ZipFile(
        temporary, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            target.writestr(item, source.read(item.filename))
        target.writestr(part_name, content)
    shutil.move(temporary, path)


def clean(path: Path) -> Path:
    """A workbook that uses every supported function and reference style.

    This is the positive control. If preflight ever rejects this, the grammar
    and the validator have drifted apart.
    """
    workbook = openpyxl.Workbook()

    assumptions = workbook.active
    assumptions.title = "Assumptions"
    assumptions["A1"] = "Growth rate"
    assumptions["B1"] = 0.08
    assumptions["A2"] = "Margin"
    assumptions["B2"] = 0.35

    model = workbook.create_sheet("Model")
    for row, value in enumerate([100, 120, 140, 160], start=1):
        model[f"A{row}"] = value

    model["B1"] = "=SUM(A1:A4)"
    model["B2"] = "=AVERAGE(A1:A4)"
    model["B3"] = "=MIN(A1:A4)"
    model["B4"] = "=MAX(A1:A4)"
    model["B5"] = "=ROUND(B2,2)"
    model["B6"] = "=ABS(A1-A4)"
    model["B7"] = '=SUMIF(A1:A4,">120")'
    model["B8"] = '=IF(A1>0,"up (good)","down")'  # a paren inside a string

    model["C1"] = "=Assumptions!B1"  # cross sheet
    model["C2"] = "=$A$1"  # absolute
    model["C3"] = "=A$1"  # mixed, locked row
    model["C4"] = "=$A1"  # mixed, locked column
    model["C5"] = "=A1*10%"  # percentage
    model["C6"] = "=-A1"  # unary minus
    model["C7"] = "=(A1+A2)*Assumptions!$B$1"
    model["C8"] = "=A1+A2-A3*A4/2"

    workbook.save(path)
    return path


def vba(path: Path) -> Path:
    """A workbook carrying a VBA project part."""
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = 1
    workbook.save(path)
    _inject_zip_part(path, "xl/vbaProject.bin", b"\x00fake vba project\x00")
    return path


def external_link_part(path: Path) -> Path:
    """A workbook declaring a link to another workbook, structurally."""
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = 1
    workbook.save(path)
    _inject_zip_part(
        path,
        "xl/externalLinks/externalLink1.xml",
        b'<?xml version="1.0"?><externalLink/>',
    )
    return path


def external_link_formula(path: Path) -> Path:
    """A workbook whose formula points into another workbook."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = 1
    sheet["A2"] = "='[1]Budget'!B4"
    workbook.save(path)
    return path


def array_formula(path: Path) -> Path:
    """A workbook with a legacy CSE array formula."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row, value in enumerate([1, 2, 3], start=1):
        sheet[f"A{row}"] = value
    sheet["B1"] = ArrayFormula(ref="B1:B3", text="=A1:A3*2")
    workbook.save(path)
    return path


def circular(path: Path) -> Path:
    """Two cells that reference each other."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "=A2+1"
    sheet["A2"] = "=A1+1"
    workbook.save(path)
    return path


def circular_via_range(path: Path) -> Path:
    """A total that sits inside the range it totals.

    Worth its own fixture: expanding the range naively is how this gets
    missed, and it is a common real error.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in range(1, 5):
        sheet[f"A{row}"] = row * 10
    sheet["A5"] = "=SUM(A1:A10)"
    workbook.save(path)
    return path


def circular_cross_sheet(path: Path) -> Path:
    """A loop that only closes when you follow a cross sheet reference."""
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "One"
    second = workbook.create_sheet("Two")
    first["A1"] = "=Two!A1+1"
    second["A1"] = "=One!A1+1"
    workbook.save(path)
    return path


def deep_chain(path: Path) -> Path:
    """A long chain and a diamond, neither of which is a cycle.

    The negative control for circular detection. A validator that flags depth
    as a loop would reject every real forecast model, so this has to pass.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Chain"
    other = workbook.create_sheet("Other")

    sheet["A1"] = 10
    for row in range(2, 12):  # A2 = A1 + 1, ten hops deep
        sheet[f"A{row}"] = f"=A{row - 1}+1"

    # Diamond: two cells read the same precedent, then a third reads both.
    sheet["B1"] = "=A11*2"
    sheet["C1"] = "=A11*3"
    sheet["D1"] = "=B1+C1"

    # A cross sheet hop that comes back to a different cell, not a loop.
    other["A1"] = "=Chain!D1+1"
    sheet["E1"] = "=Other!A1*2"

    workbook.save(path)
    return path


def defined_name(path: Path) -> Path:
    """A workbook where Q1 is a defined name, not cell Q1.

    `=Q1*2` is the trap: it reads as a perfectly ordinary cell reference.
    A print area is set as well, since that is a built in `_xlnm.` name and
    must not cause a rejection on its own.
    """
    from openpyxl.workbook.defined_name import DefinedName

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["B5"] = 0.2
    sheet["A1"] = "=Q1*2"
    sheet.print_area = "A1:D10"
    workbook.defined_names.add(DefinedName("Q1", attr_text=f"{sheet.title}!$B$5"))
    workbook.save(path)
    return path


def print_area_only(path: Path) -> Path:
    """A print area and nothing else. Must still be accepted."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = 1
    sheet["A2"] = "=A1*2"
    sheet.print_area = "A1:D10"
    workbook.save(path)
    return path


def unparseable_formula(path: Path) -> Path:
    """A formula that is inside no grammar at all.

    Not an unsupported function, just malformed. It gets its own reason code
    because "we cannot read this" and "we do not support this function" are
    different problems for the user.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = 1
    sheet["A2"] = "=A1+"
    workbook.save(path)
    return path


def copied_formulas(path: Path) -> Path:
    """Rows of copied formulas, one of them filled from the wrong origin.

    Every row here is a block that should normalise to a single R1C1 token.
    Model!F17 is the planted break: it was dragged from D17 instead of E17,
    which is the fill handle error that motivates the whole normaliser.

    Row 17 sits twelve rows below the growth assumptions in row 5, so the
    correct token is the one in docs/ARCHITECTURE.md section 2.
    """
    workbook = openpyxl.Workbook()

    assumptions = workbook.active
    assumptions.title = "Assumptions"
    assumptions["B2"] = 0.35

    model = workbook.create_sheet("Model")
    model["B2"] = 0.4
    for column in "BCDEFGH":
        model[f"{column}5"] = 0.03

    # Horizontal copy, relative references. F17 is the planted break.
    model["B17"] = 1000
    for column, previous in zip("CDEFGH", "BCDEFG"):
        model[f"{column}17"] = f"={previous}17*(1+{previous}5)"
    model["F17"] = "=D17*(1+E5)"

    # Absolute reference, unchanged by the copy.
    for column in "CDEFGH":
        model[f"{column}19"] = f"={column}17*$B$2"

    # Mixed, locked row.
    for column in "CDEFGH":
        model[f"{column}21"] = f"={column}17*{column}$5"

    # Mixed, locked column.
    for column in "CDEFGH":
        model[f"{column}23"] = f"=$B17*{column}17"

    # Cross sheet, which keeps its qualifier through normalisation.
    for column in "CDEFGH":
        model[f"{column}25"] = f"=Assumptions!$B$2*{column}17"

    # Vertical copy, to prove the same holds down a column.
    for row in range(10, 15):
        model[f"I{row}"] = row
        model[f"J{row}"] = f"=I{row}*2"

    workbook.save(path)
    return path


def unsupported_function(path: Path) -> Path:
    """A workbook using a function outside the grammar."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row, value in enumerate([1, 2, 3], start=1):
        sheet[f"A{row}"] = value
    sheet["B1"] = "=VLOOKUP(A1,A1:A3,1,FALSE)"
    workbook.save(path)
    return path


def unsupported_function_lookalike(path: Path) -> Path:
    """A function name that is also a valid cell reference.

    LOG10 parses as column LOG row 10. It has to be read as a function call,
    not silently accepted as a reference.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = 100
    sheet["B1"] = "=LOG10(A1)"
    workbook.save(path)
    return path


def unsupported_function_nested(path: Path) -> Path:
    """An unsupported function hidden inside a supported one."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = 1
    sheet["B1"] = "=SUM(A1,SQRT(A1))"
    workbook.save(path)
    return path


BUILDERS = {
    "clean": clean,
    "vba": vba,
    "external_link_part": external_link_part,
    "external_link_formula": external_link_formula,
    "array_formula": array_formula,
    "circular": circular,
    "circular_via_range": circular_via_range,
    "circular_cross_sheet": circular_cross_sheet,
    "deep_chain": deep_chain,
    "copied_formulas": copied_formulas,
    "defined_name": defined_name,
    "print_area_only": print_area_only,
    "unparseable_formula": unparseable_formula,
    "unsupported_function": unsupported_function,
    "unsupported_function_lookalike": unsupported_function_lookalike,
    "unsupported_function_nested": unsupported_function_nested,
}


def build_all(directory: Path) -> dict[str, Path]:
    directory.mkdir(parents=True, exist_ok=True)
    return {
        name: builder(directory / f"{name}.xlsx") for name, builder in BUILDERS.items()
    }


if __name__ == "__main__":
    target = Path(sys.argv[1] if len(sys.argv) > 1 else "fixtures")
    for name, written in sorted(build_all(target).items()):
        print(f"{name:34} {written}")
