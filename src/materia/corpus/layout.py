"""Where everything sits in a generated workbook.

Separated so the formula writer and the independent value calculation can
agree on addresses while disagreeing about nothing else. See generate.py for
why those two are kept apart.
"""

from openpyxl.utils import get_column_letter

MONTHS = 24
FIRST_MONTH_COLUMN = 3  # column C
TOTAL_COLUMN = FIRST_MONTH_COLUMN + MONTHS  # column AA


def month_column(month: int) -> str:
    """Column letter for a month, 1 based. Month 1 is C, month 24 is Z."""
    return get_column_letter(FIRST_MONTH_COLUMN + month - 1)


TOTAL = get_column_letter(TOTAL_COLUMN)

# Assumptions sheet, row per driver.
ASSUMPTION_ROWS = {
    "opening_customers": 4,
    "new_customer_rate": 5,
    "churn_rate": 6,
    "opening_arpu": 7,
    "arpu_uplift": 8,
    "cogs_rate": 11,
    "opening_headcount": 12,
    "hires_per_month": 13,
    "average_salary": 14,
    "payroll_tax_rate": 15,
    "marketing_rate": 16,
    "monthly_overhead": 17,
    "overhead_inflation": 18,
    "ebitda_multiple": 21,
    "net_debt": 22,
}

# Revenue sheet
REVENUE_ROWS = {
    "month": 3,
    "opening_customers": 5,
    "new_customers": 6,
    "churned_customers": 7,
    "closing_customers": 9,
    "arpu": 11,
    "average_customers": 13,
    "revenue": 15,
    "cumulative_revenue": 17,
}

# Costs sheet
COST_ROWS = {
    "month": 3,
    "headcount": 5,
    "salary": 6,
    "payroll_tax": 7,
    "staff_total": 8,
    "cogs": 10,
    "marketing": 11,
    "overhead": 12,
    "total_costs": 14,
}

# P&L sheet
PL_ROWS = {
    "month": 3,
    "revenue": 5,
    "cogs": 6,
    "gross_profit": 7,
    "gross_margin": 8,
    "staff": 10,
    "marketing": 11,
    "overhead": 12,
    "total_opex": 13,
    "ebitda": 15,
    "ebitda_margin": 16,
    "cumulative_ebitda": 18,
}

PL_TOTAL_ROWS = [
    PL_ROWS["revenue"],
    PL_ROWS["cogs"],
    PL_ROWS["gross_profit"],
    PL_ROWS["staff"],
    PL_ROWS["marketing"],
    PL_ROWS["overhead"],
    PL_ROWS["total_opex"],
    PL_ROWS["ebitda"],
]

# Valuation sheet, row per line
VALUATION_ROWS = {
    "total_revenue": 3,
    "total_ebitda": 4,
    "exit_ebitda": 5,
    "multiple": 6,
    "enterprise_value": 7,
    "net_debt": 8,
    "equity_value": 9,
    "average_monthly_ebitda": 11,
    "best_month": 12,
    "worst_month": 13,
    "positive_months": 14,
    "swing": 15,
}

ASSUMPTIONS_SHEET = "Assumptions"
REVENUE_SHEET = "Revenue"
COSTS_SHEET = "Costs"
PL_SHEET = "P&L"
VALUATION_SHEET = "Valuation"

# What materiality is measured against. docs/EVALUATION.md section 2.
DECLARED_OUTPUTS = [
    f"{PL_SHEET}!{TOTAL}{PL_ROWS['ebitda']}",
    f"{VALUATION_SHEET}!B{VALUATION_ROWS['enterprise_value']}",
]

# The last twelve months, for the exit run rate.
EXIT_FIRST_MONTH = MONTHS - 11
