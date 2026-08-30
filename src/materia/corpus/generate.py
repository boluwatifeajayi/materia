"""Corpus workbook generator.

Produces a realistic three statement forecast: assumptions, a revenue build
over 24 monthly columns, a cost build with headcount driven staff costs, a
P&L rolling to EBITDA, and a valuation applying a multiple. See
docs/EVALUATION.md section 2.

Deterministic from a seed. The same seed produces a byte identical file,
which is what `make corpus-check` compares against committed checksums.

Two independent calculations
----------------------------

The workbook is built twice. `_write_formulas` writes Excel formulas.
`compute_values` computes the same model again in plain Python, as a month by
month loop, without touching the parser or the evaluator. The second result is
written into the file as each formula's cached value.

That is deliberate. A generated workbook has no Excel written values in it, so
checking the recompute engine against a file it produced itself would be
checking the engine against the engine. Two separate implementations of the
same model, one as formulas and one as loops, disagreeing is a real signal.

They share `_excel_round`, because rounding half away from zero is a primitive
with its own tests rather than part of the model logic.
"""

from __future__ import annotations

import io
import random
import re
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.workbook.workbook import Workbook

from materia.corpus.layout import (
    ASSUMPTION_ROWS,
    # Re-exported through materia.corpus.__init__ for callers that want the
    # declared outputs without reaching into layout. Unused in this module by
    # design, so a linter will call it dead. It is not.
    DECLARED_OUTPUTS,  # noqa: F401
    ASSUMPTIONS_SHEET,
    COST_ROWS,
    COSTS_SHEET,
    EXIT_FIRST_MONTH,
    MONTHS,
    PL_ROWS,
    PL_SHEET,
    PL_TOTAL_ROWS,
    REVENUE_ROWS,
    REVENUE_SHEET,
    TOTAL,
    VALUATION_ROWS,
    VALUATION_SHEET,
    month_column,
)
from materia.recompute import _excel_round as excel_round

# A fixed timestamp for every part in the zip and for the document properties.
# Without it the same seed produces a different file every second.
FIXED_TIMESTAMP = (1980, 1, 1, 0, 0, 0)
FIXED_DATETIME = datetime(*FIXED_TIMESTAMP)


@dataclass(frozen=True)
class LegitimateBreak:
    """A cell that differs from its peers on purpose.

    Real models are full of these and they are the reason a structural
    detector alone is not useful. They are recorded in the manifest so the
    evaluator can check that a system reported none of them, which is the
    behaviour the whole project claims.
    """

    kind: str
    cells: tuple[str, ...]
    why: str


def legitimate_breaks_for(values: ModelValues) -> tuple[
    list[LegitimateBreak], dict[str, float], dict[str, str]
]:
    """The three breaks C10 carries, plus the cells that implement them.

    Returns the breaks to record, the cells to hardcode, and the comments to
    attach. See docs/EVALUATION.md section 2: C10 is the workbook that breaks
    naive tools, and every one of these is correct model building.
    """
    actuals = tuple(
        f"{REVENUE_SHEET}!{month_column(month)}{REVENUE_ROWS['revenue']}"
        for month in (1, 2, 3)
    )
    first_period = (
        f"{REVENUE_SHEET}!{month_column(1)}{REVENUE_ROWS['opening_customers']}",
        f"{REVENUE_SHEET}!{month_column(1)}{REVENUE_ROWS['arpu']}",
        f"{COSTS_SHEET}!{month_column(1)}{COST_ROWS['headcount']}",
    )
    override_cell = f"{COSTS_SHEET}!{month_column(7)}{COST_ROWS['overhead']}"
    override_note = (
        "One off office move approved by the board in month 7. Held at this "
        "figure on purpose, do not restore the inflation formula."
    )

    breaks = [
        LegitimateBreak(
            kind="hardcoded_actuals",
            cells=actuals,
            why="Months 1 to 3 are reported actuals, not forecast, so they are "
            "entered rather than calculated.",
        ),
        LegitimateBreak(
            kind="first_period",
            cells=first_period,
            why="Month 1 has no prior period to roll forward from, so it reads "
            "an assumption where later months read the month before.",
        ),
        LegitimateBreak(
            kind="manual_override",
            cells=(override_cell,),
            why=override_note,
        ),
    ]

    hardcoded = {cell: values.cells[cell] for cell in actuals}
    hardcoded[override_cell] = excel_round(values.cells[override_cell] * 1.4, 0)
    return breaks, hardcoded, {override_cell: override_note}


class MissingComputedValue(Exception):
    """A formula cell was written with no value from the Python calculation."""


@dataclass(frozen=True)
class Assumptions:
    """The drivers a workbook is built from, drawn from a seed."""

    opening_customers: int
    new_customer_rate: float
    churn_rate: float
    opening_arpu: float
    arpu_uplift: float
    cogs_rate: float
    opening_headcount: int
    hires_per_month: int
    average_salary: int
    payroll_tax_rate: float
    marketing_rate: float
    monthly_overhead: int
    overhead_inflation: float
    ebitda_multiple: float
    net_debt: int

    @classmethod
    def from_seed(cls, seed: int) -> "Assumptions":
        """Draw a set of drivers.

        Headcount and overhead are scaled to the size of the business rather
        than drawn independently. Independent draws let a small customer base
        get a large salary bill, which produced loss making forecasts on some
        seeds. That is not how a company is staffed, and a workbook nobody
        would put a multiple on cannot test a materiality gate: with a
        negative enterprise value, a percentage change in it means very little.
        """
        source = random.Random(seed)

        opening_customers = source.randrange(6_000, 12_000, 100)
        opening_arpu = round(source.uniform(60, 120), 2)
        monthly_revenue = opening_customers * opening_arpu

        return cls(
            opening_customers=opening_customers,
            new_customer_rate=round(source.uniform(0.03, 0.09), 4),
            churn_rate=round(source.uniform(0.010, 0.030), 4),
            opening_arpu=opening_arpu,
            arpu_uplift=round(source.uniform(0.002, 0.010), 4),
            cogs_rate=round(source.uniform(0.18, 0.28), 4),
            # Roughly one person per 25k to 45k of monthly revenue.
            opening_headcount=max(
                8, round(monthly_revenue / source.uniform(25_000, 45_000))
            ),
            hires_per_month=source.randint(1, 3),
            average_salary=source.randrange(55_000, 85_000, 500),
            payroll_tax_rate=round(source.uniform(0.10, 0.16), 4),
            marketing_rate=round(source.uniform(0.08, 0.14), 4),
            # Three to six percent of monthly revenue, to the nearest 500.
            monthly_overhead=round(
                monthly_revenue * source.uniform(0.03, 0.06) / 500
            ) * 500,
            overhead_inflation=round(source.uniform(0.002, 0.008), 4),
            ebitda_multiple=round(source.uniform(6, 14), 1),
            net_debt=source.randrange(0, 5_000_000, 50_000),
        )


@dataclass
class ModelValues:
    """The independently computed result, by address.

    `overrides` are cells whose value is fixed rather than computed, which is
    how C10 carries a hardcoded actuals row and a manual override. `set`
    returns the value that was actually stored, so a caller that feeds a
    result into the next month picks the override up rather than the value it
    would otherwise have computed.
    """

    overrides: dict[str, float] = field(default_factory=dict)
    cells: dict[str, float] = field(default_factory=dict)

    def set(self, sheet: str, coordinate: str, value: float) -> float:
        address = f"{sheet}!{coordinate}"
        stored = float(self.overrides.get(address, value))
        self.cells[address] = stored
        return stored


def compute_values(
    assumptions: Assumptions, overrides: dict[str, float] | None = None
) -> ModelValues:
    """Compute the whole model again, in plain Python.

    A month by month loop with no AST and no evaluator. This is the
    independent half of the cross check described in the module docstring, so
    it deliberately does not reuse anything from the formula writer beyond the
    layout constants and the rounding rule.
    """
    values = ModelValues(overrides=dict(overrides or {}))
    a = assumptions

    opening = [0.0] * (MONTHS + 1)
    closing = [0.0] * (MONTHS + 1)
    arpu = [0.0] * (MONTHS + 1)
    revenue = [0.0] * (MONTHS + 1)
    cumulative_revenue = 0.0

    for month in range(1, MONTHS + 1):
        column = month_column(month)
        values.set(REVENUE_SHEET, f"{column}{REVENUE_ROWS['month']}", month)

        opening[month] = a.opening_customers if month == 1 else closing[month - 1]
        new = excel_round(opening[month] * a.new_customer_rate, 0)
        churned = -excel_round(opening[month] * a.churn_rate, 0)
        closing[month] = opening[month] + new + churned

        arpu[month] = (
            a.opening_arpu
            if month == 1
            else excel_round(arpu[month - 1] * (1 + a.arpu_uplift), 2)
        )
        average_customers = excel_round((opening[month] + closing[month]) / 2, 0)
        revenue[month] = excel_round(average_customers * arpu[month], 0)
        cumulative_revenue += revenue[month]

        opening[month] = values.set(
            REVENUE_SHEET, f"{column}{REVENUE_ROWS['opening_customers']}", opening[month]
        )
        values.set(REVENUE_SHEET, f"{column}{REVENUE_ROWS['new_customers']}", new)
        values.set(REVENUE_SHEET, f"{column}{REVENUE_ROWS['churned_customers']}", churned)
        closing[month] = values.set(
            REVENUE_SHEET, f"{column}{REVENUE_ROWS['closing_customers']}", closing[month]
        )
        arpu[month] = values.set(REVENUE_SHEET, f"{column}{REVENUE_ROWS['arpu']}", arpu[month])
        values.set(REVENUE_SHEET, f"{column}{REVENUE_ROWS['average_customers']}", average_customers)
        revenue[month] = values.set(
            REVENUE_SHEET, f"{column}{REVENUE_ROWS['revenue']}", revenue[month]
        )
        cumulative_revenue = values.set(
            REVENUE_SHEET, f"{column}{REVENUE_ROWS['cumulative_revenue']}", cumulative_revenue
        )

    headcount = 0.0
    overhead = 0.0
    cumulative_ebitda = 0.0
    ebitda = [0.0] * (MONTHS + 1)
    totals = {row: 0.0 for row in PL_TOTAL_ROWS}

    for month in range(1, MONTHS + 1):
        column = month_column(month)
        values.set(COSTS_SHEET, f"{column}{COST_ROWS['month']}", month)

        headcount = a.opening_headcount if month == 1 else headcount + a.hires_per_month
        salary = excel_round(headcount * a.average_salary / 12, 0)
        payroll_tax = excel_round(salary * a.payroll_tax_rate, 0)
        staff_total = salary + payroll_tax
        cogs = excel_round(revenue[month] * a.cogs_rate, 0)
        marketing = excel_round(revenue[month] * a.marketing_rate, 0)
        overhead = (
            excel_round(a.monthly_overhead, 0)
            if month == 1
            else excel_round(overhead * (1 + a.overhead_inflation), 0)
        )
        headcount = values.set(COSTS_SHEET, f"{column}{COST_ROWS['headcount']}", headcount)
        salary = values.set(COSTS_SHEET, f"{column}{COST_ROWS['salary']}", salary)
        payroll_tax = values.set(COSTS_SHEET, f"{column}{COST_ROWS['payroll_tax']}", payroll_tax)
        staff_total = values.set(
            COSTS_SHEET, f"{column}{COST_ROWS['staff_total']}", salary + payroll_tax
        )
        cogs = values.set(COSTS_SHEET, f"{column}{COST_ROWS['cogs']}", cogs)
        marketing = values.set(COSTS_SHEET, f"{column}{COST_ROWS['marketing']}", marketing)
        overhead = values.set(COSTS_SHEET, f"{column}{COST_ROWS['overhead']}", overhead)
        values.set(
            COSTS_SHEET,
            f"{column}{COST_ROWS['total_costs']}",
            staff_total + cogs + marketing + overhead,
        )

        # P&L: costs are shown negative
        pl_revenue = revenue[month]
        pl_cogs = -cogs
        gross_profit = pl_revenue + pl_cogs
        gross_margin = 0.0 if pl_revenue == 0 else excel_round(gross_profit / pl_revenue, 4)
        pl_staff = -staff_total
        pl_marketing = -marketing
        pl_overhead = -overhead
        total_opex = pl_staff + pl_marketing + pl_overhead
        ebitda[month] = gross_profit + total_opex
        ebitda_margin = 0.0 if pl_revenue == 0 else excel_round(ebitda[month] / pl_revenue, 4)
        cumulative_ebitda += ebitda[month]

        row_values = {
            PL_ROWS["month"]: month,
            PL_ROWS["revenue"]: pl_revenue,
            PL_ROWS["cogs"]: pl_cogs,
            PL_ROWS["gross_profit"]: gross_profit,
            PL_ROWS["gross_margin"]: gross_margin,
            PL_ROWS["staff"]: pl_staff,
            PL_ROWS["marketing"]: pl_marketing,
            PL_ROWS["overhead"]: pl_overhead,
            PL_ROWS["total_opex"]: total_opex,
            PL_ROWS["ebitda"]: ebitda[month],
            PL_ROWS["ebitda_margin"]: ebitda_margin,
            PL_ROWS["cumulative_ebitda"]: cumulative_ebitda,
        }
        for row, value in row_values.items():
            stored = values.set(PL_SHEET, f"{column}{row}", value)
            if row == PL_ROWS["ebitda"]:
                ebitda[month] = stored
            if row == PL_ROWS["cumulative_ebitda"]:
                cumulative_ebitda = stored
            if row in totals:
                totals[row] += stored

    for row, total in totals.items():
        values.set(PL_SHEET, f"{TOTAL}{row}", total)

    monthly_ebitda = ebitda[1 : MONTHS + 1]
    exit_ebitda = sum(ebitda[EXIT_FIRST_MONTH : MONTHS + 1])
    enterprise_value = excel_round(exit_ebitda * a.ebitda_multiple, 0)
    best = max(monthly_ebitda)
    worst = min(monthly_ebitda)

    valuation = {
        "total_revenue": cumulative_revenue,
        "total_ebitda": totals[PL_ROWS["ebitda"]],
        "exit_ebitda": exit_ebitda,
        "multiple": a.ebitda_multiple,
        "enterprise_value": enterprise_value,
        "net_debt": a.net_debt,
        "equity_value": enterprise_value - a.net_debt,
        "average_monthly_ebitda": sum(monthly_ebitda) / len(monthly_ebitda),
        "best_month": best,
        "worst_month": worst,
        "positive_months": sum(value for value in monthly_ebitda if value > 0),
        "swing": abs(best - worst),
    }
    for name, value in valuation.items():
        values.set(VALUATION_SHEET, f"B{VALUATION_ROWS[name]}", value)

    return values


# --- formula writing -------------------------------------------------------


def _assumption(name: str) -> str:
    """An absolute reference to one assumption, as a copied formula would."""
    return f"{ASSUMPTIONS_SHEET}!$B${ASSUMPTION_ROWS[name]}"


def _write_assumptions(sheet, a: Assumptions) -> None:
    sheet["A1"] = "Assumptions"
    sheet["A3"] = "Revenue drivers"
    sheet["A10"] = "Cost drivers"
    sheet["A20"] = "Valuation drivers"

    labels = {
        "opening_customers": "Opening customers",
        "new_customer_rate": "Monthly new customer rate",
        "churn_rate": "Monthly churn rate",
        "opening_arpu": "Opening ARPU",
        "arpu_uplift": "Monthly ARPU uplift",
        "cogs_rate": "COGS as share of revenue",
        "opening_headcount": "Opening headcount",
        "hires_per_month": "New hires per month",
        "average_salary": "Average salary",
        "payroll_tax_rate": "Payroll tax rate",
        "marketing_rate": "Marketing as share of revenue",
        "monthly_overhead": "Monthly overhead",
        "overhead_inflation": "Monthly overhead inflation",
        "ebitda_multiple": "EBITDA multiple",
        "net_debt": "Net debt",
    }
    for name, row in ASSUMPTION_ROWS.items():
        sheet[f"A{row}"] = labels[name]
        sheet[f"B{row}"] = getattr(a, name)


def _write_revenue(sheet) -> None:
    rows = REVENUE_ROWS
    labels = {
        "month": "Month",
        "opening_customers": "Opening customers",
        "new_customers": "New customers",
        "churned_customers": "Churned customers",
        "closing_customers": "Closing customers",
        "arpu": "ARPU",
        "average_customers": "Average customers",
        "revenue": "Revenue",
        "cumulative_revenue": "Cumulative revenue",
    }
    sheet["A1"] = "Revenue build"
    for name, row in rows.items():
        sheet[f"A{row}"] = labels[name]

    for month in range(1, MONTHS + 1):
        here = month_column(month)
        previous = month_column(month - 1) if month > 1 else None

        sheet[f"{here}{rows['month']}"] = 1 if month == 1 else f"={previous}{rows['month']}+1"
        sheet[f"{here}{rows['opening_customers']}"] = (
            f"={_assumption('opening_customers')}"
            if month == 1
            else f"={previous}{rows['closing_customers']}"
        )
        sheet[f"{here}{rows['new_customers']}"] = (
            f"=ROUND({here}{rows['opening_customers']}*{_assumption('new_customer_rate')},0)"
        )
        sheet[f"{here}{rows['churned_customers']}"] = (
            f"=-ROUND({here}{rows['opening_customers']}*{_assumption('churn_rate')},0)"
        )
        sheet[f"{here}{rows['closing_customers']}"] = (
            f"={here}{rows['opening_customers']}+{here}{rows['new_customers']}"
            f"+{here}{rows['churned_customers']}"
        )
        sheet[f"{here}{rows['arpu']}"] = (
            f"={_assumption('opening_arpu')}"
            if month == 1
            else f"=ROUND({previous}{rows['arpu']}*(1+{_assumption('arpu_uplift')}),2)"
        )
        sheet[f"{here}{rows['average_customers']}"] = (
            f"=ROUND(({here}{rows['opening_customers']}"
            f"+{here}{rows['closing_customers']})/2,0)"
        )
        sheet[f"{here}{rows['revenue']}"] = (
            f"=ROUND({here}{rows['average_customers']}*{here}{rows['arpu']},0)"
        )
        sheet[f"{here}{rows['cumulative_revenue']}"] = (
            f"={here}{rows['revenue']}"
            if month == 1
            else f"={previous}{rows['cumulative_revenue']}+{here}{rows['revenue']}"
        )


def _write_costs(sheet) -> None:
    rows = COST_ROWS
    labels = {
        "month": "Month",
        "headcount": "Headcount",
        "salary": "Salary cost",
        "payroll_tax": "Payroll tax",
        "staff_total": "Total staff cost",
        "cogs": "Cost of sales",
        "marketing": "Marketing",
        "overhead": "Overhead",
        "total_costs": "Total costs",
    }
    sheet["A1"] = "Cost build"
    for name, row in rows.items():
        sheet[f"A{row}"] = labels[name]

    for month in range(1, MONTHS + 1):
        here = month_column(month)
        previous = month_column(month - 1) if month > 1 else None
        revenue_cell = f"{REVENUE_SHEET}!{here}{REVENUE_ROWS['revenue']}"

        sheet[f"{here}{rows['month']}"] = f"={REVENUE_SHEET}!{here}{REVENUE_ROWS['month']}"
        sheet[f"{here}{rows['headcount']}"] = (
            f"={_assumption('opening_headcount')}"
            if month == 1
            else f"={previous}{rows['headcount']}+{_assumption('hires_per_month')}"
        )
        sheet[f"{here}{rows['salary']}"] = (
            f"=ROUND({here}{rows['headcount']}*{_assumption('average_salary')}/12,0)"
        )
        sheet[f"{here}{rows['payroll_tax']}"] = (
            f"=ROUND({here}{rows['salary']}*{_assumption('payroll_tax_rate')},0)"
        )
        sheet[f"{here}{rows['staff_total']}"] = (
            f"={here}{rows['salary']}+{here}{rows['payroll_tax']}"
        )
        sheet[f"{here}{rows['cogs']}"] = (
            f"=ROUND({revenue_cell}*{_assumption('cogs_rate')},0)"
        )
        sheet[f"{here}{rows['marketing']}"] = (
            f"=ROUND({revenue_cell}*{_assumption('marketing_rate')},0)"
        )
        sheet[f"{here}{rows['overhead']}"] = (
            f"=ROUND({_assumption('monthly_overhead')},0)"
            if month == 1
            else f"=ROUND({previous}{rows['overhead']}*(1+{_assumption('overhead_inflation')}),0)"
        )
        sheet[f"{here}{rows['total_costs']}"] = (
            f"={here}{rows['staff_total']}+{here}{rows['cogs']}"
            f"+{here}{rows['marketing']}+{here}{rows['overhead']}"
        )


def _write_pl(sheet) -> None:
    rows = PL_ROWS
    labels = {
        "month": "Month",
        "revenue": "Revenue",
        "cogs": "Cost of sales",
        "gross_profit": "Gross profit",
        "gross_margin": "Gross margin",
        "staff": "Staff costs",
        "marketing": "Marketing",
        "overhead": "Overhead",
        "total_opex": "Total operating costs",
        "ebitda": "EBITDA",
        "ebitda_margin": "EBITDA margin",
        "cumulative_ebitda": "Cumulative EBITDA",
    }
    sheet["A1"] = "Profit and loss"
    for name, row in rows.items():
        sheet[f"A{row}"] = labels[name]
    sheet[f"{TOTAL}3"] = "Total"

    for month in range(1, MONTHS + 1):
        here = month_column(month)
        previous = month_column(month - 1) if month > 1 else None

        sheet[f"{here}{rows['month']}"] = f"={REVENUE_SHEET}!{here}{REVENUE_ROWS['month']}"
        sheet[f"{here}{rows['revenue']}"] = f"={REVENUE_SHEET}!{here}{REVENUE_ROWS['revenue']}"
        sheet[f"{here}{rows['cogs']}"] = f"=-{COSTS_SHEET}!{here}{COST_ROWS['cogs']}"
        sheet[f"{here}{rows['gross_profit']}"] = (
            f"={here}{rows['revenue']}+{here}{rows['cogs']}"
        )
        sheet[f"{here}{rows['gross_margin']}"] = (
            f"=IF({here}{rows['revenue']}=0,0,"
            f"ROUND({here}{rows['gross_profit']}/{here}{rows['revenue']},4))"
        )
        sheet[f"{here}{rows['staff']}"] = f"=-{COSTS_SHEET}!{here}{COST_ROWS['staff_total']}"
        sheet[f"{here}{rows['marketing']}"] = f"=-{COSTS_SHEET}!{here}{COST_ROWS['marketing']}"
        sheet[f"{here}{rows['overhead']}"] = f"=-{COSTS_SHEET}!{here}{COST_ROWS['overhead']}"
        sheet[f"{here}{rows['total_opex']}"] = (
            f"={here}{rows['staff']}+{here}{rows['marketing']}+{here}{rows['overhead']}"
        )
        sheet[f"{here}{rows['ebitda']}"] = (
            f"={here}{rows['gross_profit']}+{here}{rows['total_opex']}"
        )
        sheet[f"{here}{rows['ebitda_margin']}"] = (
            f"=IF({here}{rows['revenue']}=0,0,"
            f"ROUND({here}{rows['ebitda']}/{here}{rows['revenue']},4))"
        )
        sheet[f"{here}{rows['cumulative_ebitda']}"] = (
            f"={here}{rows['ebitda']}"
            if month == 1
            else f"={previous}{rows['cumulative_ebitda']}+{here}{rows['ebitda']}"
        )

    first, last = month_column(1), month_column(MONTHS)
    for row in PL_TOTAL_ROWS:
        sheet[f"{TOTAL}{row}"] = f"=SUM({first}{row}:{last}{row})"


def _write_valuation(sheet) -> None:
    rows = VALUATION_ROWS
    labels = {
        "total_revenue": "Total revenue",
        "total_ebitda": "Total EBITDA",
        "exit_ebitda": "Exit run rate EBITDA",
        "multiple": "EBITDA multiple",
        "enterprise_value": "Enterprise value",
        "net_debt": "Net debt",
        "equity_value": "Equity value",
        "average_monthly_ebitda": "Average monthly EBITDA",
        "best_month": "Best month",
        "worst_month": "Worst month",
        "positive_months": "Sum of positive months",
        "swing": "Best to worst swing",
    }
    sheet["A1"] = "Valuation"
    for name, row in rows.items():
        sheet[f"A{row}"] = labels[name]

    first, last = month_column(1), month_column(MONTHS)
    exit_first = month_column(EXIT_FIRST_MONTH)
    ebitda_row = PL_ROWS["ebitda"]
    ebitda_range = f"'{PL_SHEET}'!{first}{ebitda_row}:{last}{ebitda_row}"

    sheet[f"B{rows['total_revenue']}"] = (
        f"={REVENUE_SHEET}!{last}{REVENUE_ROWS['cumulative_revenue']}"
    )
    sheet[f"B{rows['total_ebitda']}"] = f"='{PL_SHEET}'!{TOTAL}{ebitda_row}"
    sheet[f"B{rows['exit_ebitda']}"] = (
        f"=SUM('{PL_SHEET}'!{exit_first}{ebitda_row}:{last}{ebitda_row})"
    )
    sheet[f"B{rows['multiple']}"] = f"={_assumption('ebitda_multiple')}"
    sheet[f"B{rows['enterprise_value']}"] = (
        f"=ROUND(B{rows['exit_ebitda']}*B{rows['multiple']},0)"
    )
    sheet[f"B{rows['net_debt']}"] = f"={_assumption('net_debt')}"
    sheet[f"B{rows['equity_value']}"] = (
        f"=B{rows['enterprise_value']}-B{rows['net_debt']}"
    )
    sheet[f"B{rows['average_monthly_ebitda']}"] = f"=AVERAGE({ebitda_range})"
    sheet[f"B{rows['best_month']}"] = f"=MAX({ebitda_range})"
    sheet[f"B{rows['worst_month']}"] = f"=MIN({ebitda_range})"
    sheet[f"B{rows['positive_months']}"] = f'=SUMIF({ebitda_range},">0")'
    sheet[f"B{rows['swing']}"] = (
        f"=ABS(B{rows['best_month']}-B{rows['worst_month']})"
    )


def _apply_breaks(
    workbook: Workbook, hardcoded: dict[str, float], comments: dict[str, str]
) -> None:
    """Replace formulas with entered values, and attach the notes explaining why.

    Applied after the formulas are written rather than woven into the writers,
    so the ordinary model stays readable and the breaks are visible in one
    place.
    """
    for address, value in hardcoded.items():
        sheet, coordinate = address.split("!", 1)
        workbook[sheet][coordinate] = value
    for address, text in comments.items():
        sheet, coordinate = address.split("!", 1)
        workbook[sheet][coordinate].comment = Comment(text, "Finance")

    if hardcoded:
        revenue = workbook[REVENUE_SHEET]
        revenue["A2"] = "Status"
        for month in range(1, MONTHS + 1):
            revenue[f"{month_column(month)}2"] = "Actual" if month <= 3 else "Forecast"


def build_workbook(
    assumptions: Assumptions,
    hardcoded: dict[str, float] | None = None,
    comments: dict[str, str] | None = None,
) -> Workbook:
    """The workbook with formulas in it, before values are cached."""
    workbook = openpyxl.Workbook()

    sheet = workbook.active
    sheet.title = ASSUMPTIONS_SHEET
    _write_assumptions(sheet, assumptions)
    _write_revenue(workbook.create_sheet(REVENUE_SHEET))
    _write_costs(workbook.create_sheet(COSTS_SHEET))
    _write_pl(workbook.create_sheet(PL_SHEET))
    _write_valuation(workbook.create_sheet(VALUATION_SHEET))

    # Document properties carry a timestamp by default, which would make two
    # runs of the same seed differ.
    _apply_breaks(workbook, hardcoded or {}, comments or {})

    workbook.properties.creator = "materia"
    workbook.properties.lastModifiedBy = "materia"
    workbook.properties.created = FIXED_DATETIME
    workbook.properties.modified = FIXED_DATETIME
    return workbook


# --- deterministic saving with cached values -------------------------------

_EMPTY_VALUE = re.compile(rb"(<c r=\"([A-Z]+[0-9]+)\"[^>]*>)(<f[^>]*>.*?</f>)<v\s*/>")

# openpyxl stamps the save time into core.xml whatever the properties say, so
# it is normalised on the way out rather than set on the way in.
_XML_TIMESTAMP = re.compile(rb">\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z<")
_FIXED_XML_TIMESTAMP = b">1980-01-01T00:00:00Z<"


def _sheet_parts(archive: zipfile.ZipFile) -> dict[str, str]:
    """Map sheet name to the zip part holding it.

    Read from the workbook relationships rather than assumed from ordering,
    because writing a cached value into the wrong sheet would be silent.
    """
    workbook_xml = archive.read("xl/workbook.xml").decode()
    rels_xml = archive.read("xl/_rels/workbook.xml.rels").decode()

    # Attribute order in the relationship elements is not guaranteed, and the
    # target may be absolute or relative to xl/, so both are handled rather
    # than assumed.
    targets = {}
    for element in re.findall(r"<Relationship\b[^>]*/?>", rels_xml):
        identifier = re.search(r'Id="([^"]+)"', element)
        target = re.search(r'Target="([^"]+)"', element)
        if identifier and target:
            path = target.group(1).lstrip("/")
            targets[identifier.group(1)] = path if path.startswith("xl/") else f"xl/{path}"

    parts = {}
    for name, relation in re.findall(
        r'<sheet name="([^"]+)"[^>]*r:id="([^"]+)"', workbook_xml
    ):
        parts[name.replace("&amp;", "&")] = targets[relation]
    return parts


def _inject_values(xml: bytes, sheet: str, values: dict[str, float]) -> bytes:
    """Fill each formula cell's empty <v/> with its computed value."""

    def replace(match: re.Match) -> bytes:
        cell_open, coordinate, formula = match.groups()
        address = f"{sheet}!{coordinate.decode()}"
        if address not in values:
            # Every formula cell must have a value from the independent
            # calculation. Leaving one empty would not fail loudly: the cross
            # check in the tests would simply skip that cell and still pass,
            # which is worse than a missing value.
            raise MissingComputedValue(
                f"{address} has a formula but no independently computed value"
            )
        return cell_open + formula + f"<v>{values[address]!r}</v>".encode()

    return _EMPTY_VALUE.sub(replace, xml)


def save(workbook: Workbook, path: Path, values: ModelValues) -> Path:
    """Write the workbook with cached values and a fixed timestamp.

    Timestamps are normalised so the same seed produces a byte identical file,
    which is what `make corpus-check` relies on.
    """
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(buffer) as source, zipfile.ZipFile(
        path, "w", zipfile.ZIP_DEFLATED
    ) as target:
        parts = _sheet_parts(source)
        part_to_sheet = {part: sheet for sheet, part in parts.items()}
        for item in source.infolist():
            data = source.read(item.filename)
            sheet = part_to_sheet.get(item.filename)
            if sheet is not None:
                data = _inject_values(data, sheet, values.cells)
            elif item.filename == "docProps/core.xml":
                data = _XML_TIMESTAMP.sub(_FIXED_XML_TIMESTAMP, data)
            info = zipfile.ZipInfo(item.filename, date_time=FIXED_TIMESTAMP)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = item.external_attr
            target.writestr(info, data)
    return path


def generate(
    path: str | Path, seed: int, legitimate_breaks: bool = False
) -> tuple[Path, list[LegitimateBreak]]:
    """Generate one corpus workbook. Same seed, byte identical file.

    With `legitimate_breaks`, the workbook gains the three deliberate pattern
    breaks C10 needs. It stays a clean control: nothing in it is an error.
    """
    assumptions = Assumptions.from_seed(seed)

    baseline = compute_values(assumptions)
    breaks: list[LegitimateBreak] = []
    hardcoded: dict[str, float] = {}
    comments: dict[str, str] = {}
    if legitimate_breaks:
        breaks, hardcoded, comments = legitimate_breaks_for(baseline)

    values = compute_values(assumptions, hardcoded)
    workbook = build_workbook(assumptions, hardcoded, comments)
    return save(workbook, Path(path), values), breaks
