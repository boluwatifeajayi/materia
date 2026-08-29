"""Formula reading and R1C1 normalisation.

openpyxl in formula mode gives raw A1 strings. A row of copied formulas looks
like twelve different strings in A1 and one identical string in R1C1. Peer
group comparison is only possible after normalisation, so this happens before
anything else. See docs/ARCHITECTURE.md section 2.

    G17 = F17*(1+F5)  ->  RC[-1]*(1+R[-12]C[-1])
    H17 = G17*(1+G5)  ->  RC[-1]*(1+R[-12]C[-1])

A cell in that row that does not normalise to the same token is the signal
every detector is built on.

This module owns what a reference looks like. Preflight imports from here so
the definition lives in one place.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

# Excel escapes a double quote inside a string literal by doubling it.
STRING_LITERAL = re.compile(r'"(?:[^"]|"")*"')

_SHEET = r"(?:'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!"

# The trailing (?![0-9]) matters more than it looks. Without it the engine
# backtracks: in LOG10( the greedy digit match takes LOG10, fails the "not a
# function call" lookahead below, then retries as LOG1 and succeeds, leaving a
# stray 0 behind. Refusing a partial digit run closes that off.
_CELL = r"\$?[A-Za-z]{1,3}\$?[0-9]{1,7}(?![0-9])"

# One cell reference, with its parts named so it can be rewritten.
SINGLE_REFERENCE = re.compile(
    rf"(?:(?P<sheet>{_SHEET[:-1]})!)?"
    r"(?P<column_absolute>\$?)(?P<column>[A-Za-z]{1,3})"
    r"(?P<row_absolute>\$?)(?P<row>[0-9]{1,7})"
)

# A cell or a range. The lookbehind stops it matching the tail of an
# identifier. The lookahead stops it swallowing a function name that is also a
# valid cell reference, such as LOG10( , where LOG is a real column.
REFERENCE = re.compile(
    rf"(?<![A-Za-z0-9_.])(?:{_SHEET})?{_CELL}(?::(?:{_SHEET})?{_CELL})?(?!\s*\()"
)


class InvalidReference(ValueError):
    """Raised when text that should be a cell reference is not one."""


@dataclass(frozen=True)
class Reference:
    """One cell reference, decomposed.

    `row` and `column` are always 1 based absolute positions. The two
    `absolute_*` flags record whether the author wrote a dollar sign, which is
    what decides between an offset and a fixed index in R1C1.
    """

    column: int
    row: int
    absolute_column: bool = False
    absolute_row: bool = False
    sheet: str | None = None

    @property
    def a1(self) -> str:
        column = ("$" if self.absolute_column else "") + get_column_letter(self.column)
        row = ("$" if self.absolute_row else "") + str(self.row)
        qualifier = f"{self.sheet}!" if self.sheet else ""
        return f"{qualifier}{column}{row}"

    def to_r1c1(self, origin_row: int, origin_column: int) -> str:
        """Render relative to the cell the formula sits in.

        An absolute part becomes a fixed 1 based index. A relative part becomes
        a bracketed offset from the origin, or nothing at all when the offset
        is zero, which is what makes `RC[-1]` mean "the cell to my left".
        """
        row = _axis(self.row, origin_row, self.absolute_row)
        column = _axis(self.column, origin_column, self.absolute_column)
        qualifier = f"{self.sheet}!" if self.sheet else ""
        return f"{qualifier}R{row}C{column}"


def _axis(index: int, origin: int, absolute: bool) -> str:
    if absolute:
        return str(index)
    offset = index - origin
    return f"[{offset}]" if offset else ""


def parse_reference(text: str) -> Reference:
    """Turn "Sheet1!$B$5" into a Reference. Raises on anything else."""
    match = SINGLE_REFERENCE.fullmatch(text.strip())
    if match is None:
        raise InvalidReference(text)
    return Reference(
        column=column_index_from_string(match["column"].upper()),
        row=int(match["row"]),
        absolute_column=bool(match["column_absolute"]),
        absolute_row=bool(match["row_absolute"]),
        sheet=match["sheet"],
    )


def strip_string_literals(formula: str) -> str:
    """Blank out string literals so their contents are never scanned."""
    return STRING_LITERAL.sub('""', formula)


def references_in(formula: str) -> list[Reference]:
    """Every cell reference in a formula, ranges yielding both endpoints."""
    found = []
    for match in REFERENCE.finditer(strip_string_literals(formula)):
        for part in match.group(0).split(":"):
            found.append(parse_reference(part))
    return found


def to_r1c1(reference: str, *, row: int, column: int) -> str:
    """Normalise one reference, or one range, against an origin cell."""
    return ":".join(
        parse_reference(part).to_r1c1(row, column) for part in reference.split(":")
    )


def normalise(formula: str, *, row: int, column: int) -> str:
    """Normalise a whole formula into its R1C1 peer comparison token.

    Three things happen, and only outside string literals: references become
    R1C1, insignificant whitespace goes, and the rest is uppercased so a hand
    typed `sum(` matches Excel's `SUM(`. Literals are passed through exactly,
    because two formulas carrying different text are genuinely different.

    The leading `=` is dropped. The result is a comparison token, not a
    formula.
    """
    body = formula[1:] if formula.startswith("=") else formula

    pieces: list[str] = []
    position = 0
    pattern = re.compile(f"(?P<literal>{STRING_LITERAL.pattern})|(?P<ref>{REFERENCE.pattern})")

    for match in pattern.finditer(body):
        pieces.append(_normalise_gap(body[position : match.start()]))
        if match["literal"] is not None:
            pieces.append(match["literal"])
        else:
            pieces.append(to_r1c1(match["ref"], row=row, column=column))
        position = match.end()

    pieces.append(_normalise_gap(body[position:]))
    return "".join(pieces)


def _normalise_gap(text: str) -> str:
    """Everything that is not a literal or a reference: operators, numbers,
    function names, punctuation."""
    return re.sub(r"\s+", "", text).upper()


# A range wider than this is refused rather than enumerated. Real models do
# not aggregate over a million cells, and silently taking minutes to walk one
# would be worse than saying no.
MAX_RANGE_CELLS = 65_536


class RangeTooLarge(ValueError):
    """A range covers more cells than Materia will enumerate."""


def normalise_address(address: str) -> str:
    """Put an address into the one form used as a key everywhere.

    Sheet quotes and dollar signs are presentation, not identity, so
    `\'My Sheet\'!$B$5` and `My Sheet!b5` are the same cell.
    """
    sheet, _, coordinate = address.rpartition("!")
    return f"{sheet.strip(chr(39))}!{coordinate.replace('$', '').upper()}"


def cell_address(reference: Reference, default_sheet: str) -> str:
    """Resolve one reference to an absolute address."""
    sheet = (reference.sheet or default_sheet).strip("'")
    return f"{sheet}!{get_column_letter(reference.column)}{reference.row}"


def range_addresses(
    start: Reference, end: Reference, default_sheet: str
) -> Iterator[str]:
    """Every address in a rectangle, row then column order.

    Empty cells are included, because SUMIF needs to line a criteria range up
    with a sum range of the same shape.
    """
    sheet = (start.sheet or default_sheet).strip("'")
    first_row, last_row = sorted((start.row, end.row))
    first_column, last_column = sorted((start.column, end.column))

    size = (last_row - first_row + 1) * (last_column - first_column + 1)
    if size > MAX_RANGE_CELLS:
        raise RangeTooLarge(
            f"range covers {size} cells, above the {MAX_RANGE_CELLS} limit"
        )

    for row in range(first_row, last_row + 1):
        for column in range(first_column, last_column + 1):
            yield f"{sheet}!{get_column_letter(column)}{row}"


@dataclass(frozen=True)
class FormulaCell:
    """One formula, in both the form the author wrote and the comparable one."""

    sheet: str
    coordinate: str
    row: int
    column: int
    formula: str
    r1c1: str

    @property
    def address(self) -> str:
        return f"{self.sheet}!{self.coordinate}"


def read_formulas(path: str | Path) -> list[FormulaCell]:
    """Read every formula in a workbook, normalised, in sheet order.

    The workbook is opened read only and is never written. Run preflight
    first: this assumes the formulas are inside the supported grammar.
    """
    workbook = openpyxl.load_workbook(Path(path), read_only=True, data_only=False)
    try:
        cells = []
        for sheet_name in workbook.sheetnames:
            for row in workbook[sheet_name].iter_rows():
                for cell in row:
                    value = cell.value
                    if not (isinstance(value, str) and value.startswith("=")):
                        continue
                    cells.append(
                        FormulaCell(
                            sheet=sheet_name,
                            coordinate=cell.coordinate,
                            row=cell.row,
                            column=cell.column,
                            formula=value,
                            r1c1=normalise(value, row=cell.row, column=cell.column),
                        )
                    )
        return cells
    finally:
        workbook.close()
