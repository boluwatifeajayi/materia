"""Preflight validator.

Materia refuses workbooks it cannot reason about faithfully, rather than
guessing. A tool that mis-evaluates a formula it does not understand produces
confident wrong impact numbers, which is worse than producing nothing. See
docs/ARCHITECTURE.md section 1 and README.md section 6.

The workbook is opened read only and is never written.
"""

from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path

import networkx as nx
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.formula import ArrayFormula

# README.md section 6. Do not extend without changing that section and the
# recompute engine in the same commit: every function here has to be
# faithfully evaluable by src/materia/recompute.py.
SUPPORTED_FUNCTIONS = frozenset(
    {"SUM", "AVERAGE", "MIN", "MAX", "IF", "ROUND", "ABS", "SUMIF"}
)


class Reason(str, Enum):
    """Named rejection reasons. The user always gets one of these."""

    VBA_PRESENT = "VBA_PRESENT"
    EXTERNAL_LINK = "EXTERNAL_LINK"
    ARRAY_FORMULA = "ARRAY_FORMULA"
    CIRCULAR_REFERENCE = "CIRCULAR_REFERENCE"
    UNSUPPORTED_FUNCTION = "UNSUPPORTED_FUNCTION"


class PreflightRejected(Exception):
    """Raised when a workbook is outside what Materia can evaluate faithfully.

    Carries a machine readable `code` and a message that names where the
    problem is, so the user can go and look at it.
    """

    def __init__(
        self,
        reason: Reason,
        detail: str,
        *,
        location: str | None = None,
        function: str | None = None,
    ) -> None:
        self.reason = reason
        self.detail = detail
        self.location = location
        self.function = function
        super().__init__(self.message)

    @property
    def code(self) -> str:
        if self.reason is Reason.UNSUPPORTED_FUNCTION:
            return f"{Reason.UNSUPPORTED_FUNCTION.value}({self.function})"
        return self.reason.value

    @property
    def message(self) -> str:
        where = f" at {self.location}" if self.location else ""
        return f"{self.code}{where}: {self.detail}"


@dataclass
class PreflightReport:
    """What a workbook contains, once it has been accepted.

    The formula count is the first line of the funnel in README section 4.
    """

    path: Path
    sheet_names: list[str] = field(default_factory=list)
    formula_count: int = 0
    value_cell_count: int = 0


# --- formula scanning helpers ---------------------------------------------

# Excel escapes a double quote inside a string literal by doubling it.
_STRING_LITERAL = re.compile(r'"(?:[^"]|"")*"')

# An identifier immediately followed by "(" is a function call. The name may
# carry Excel's _xlfn. prefix, which marks a function newer than the file
# format and is therefore never in our grammar.
_FUNCTION_CALL = re.compile(r"(?<![A-Za-z0-9_.])([A-Za-z_][A-Za-z0-9_.]*)\s*\(")

_SHEET_QUALIFIER = r"(?:(?:'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!)?"
_CELL = r"\$?[A-Za-z]{1,3}\$?[0-9]{1,7}"

# A cell or range reference. The lookbehind stops it matching the tail of an
# identifier, and the lookahead stops it swallowing a function name that is
# also a valid cell reference, such as LOG10( .
_REFERENCE = re.compile(
    rf"(?<![A-Za-z0-9_.]){_SHEET_QUALIFIER}{_CELL}(?::{_SHEET_QUALIFIER}?{_CELL})?(?!\s*\()"
)

# A reference into another workbook: [1]Sheet1!A1, '[1]Sheet 1'!A1, or a
# reference carrying a full path.
_EXTERNAL_REFERENCE = re.compile(r"(?:'[^']*\[[^\]]+\][^']*'|\[[^\]]+\])")


def _strip_string_literals(formula: str) -> str:
    """Blank out string literals so their contents are never scanned."""
    return _STRING_LITERAL.sub('""', formula)


def _functions_in(formula: str) -> list[str]:
    return _FUNCTION_CALL.findall(_strip_string_literals(formula))


def _split_reference(ref: str) -> tuple[str | None, str]:
    """Split "Sheet1!A1" into ("Sheet1", "A1"). Unqualified gives (None, ref)."""
    if "!" not in ref:
        return None, ref.replace("$", "")
    sheet, _, cell = ref.rpartition("!")
    return sheet.strip("'"), cell.replace("$", "")


# --- individual checks -----------------------------------------------------


def _check_container(path: Path) -> None:
    """Check the parts inside the .xlsx zip, before parsing any formula.

    VBA and external links are declared structurally, so they are visible here
    without reading a single cell.
    """
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
    except zipfile.BadZipFile as exc:
        # Not a PreflightRejected. That means "a real workbook containing
        # something we cannot evaluate", and the reason codes are the five in
        # docs/ARCHITECTURE.md section 1. A file that is not an .xlsx at all
        # is a different problem and mislabelling it would be misleading.
        raise ValueError(f"{path} is not a readable .xlsx file") from exc

    if any(name.startswith("xl/vbaProject") for name in names):
        raise PreflightRejected(
            Reason.VBA_PRESENT,
            "the workbook contains a VBA project. Macros can change values in "
            "ways Materia cannot see, so any impact figure would be unsound.",
        )

    if any(name.startswith("xl/externalLinks/") for name in names):
        raise PreflightRejected(
            Reason.EXTERNAL_LINK,
            "the workbook links to another workbook. Materia cannot read the "
            "other file, so it cannot recompute this one.",
        )


def _check_formulas(
    workbook: openpyxl.Workbook,
) -> tuple[dict[str, str], int]:
    """Scan every cell once.

    Returns the formulas by qualified address, and a count of non formula
    cells that hold a value. Raises on the first array formula, external
    reference or unsupported function, scanning in sheet, row, column order so
    the reported location is stable across runs.
    """
    formulas: dict[str, str] = {}
    value_cells = 0

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue

                location = f"{sheet_name}!{cell.coordinate}"

                if isinstance(value, ArrayFormula):
                    raise PreflightRejected(
                        Reason.ARRAY_FORMULA,
                        "array formulas spill across a range of cells, which "
                        "the recompute engine does not model.",
                        location=location,
                    )

                if not (isinstance(value, str) and value.startswith("=")):
                    value_cells += 1
                    continue

                formula = value
                scannable = _strip_string_literals(formula)

                if _EXTERNAL_REFERENCE.search(scannable):
                    raise PreflightRejected(
                        Reason.EXTERNAL_LINK,
                        f"{formula} refers to another workbook, which Materia "
                        "cannot read.",
                        location=location,
                    )

                for name in _functions_in(formula):
                    bare = name.split(".")[-1].upper()
                    if bare not in SUPPORTED_FUNCTIONS:
                        raise PreflightRejected(
                            Reason.UNSUPPORTED_FUNCTION,
                            f"{formula} uses {bare}, which is outside the "
                            "supported grammar in README section 6.",
                            location=location,
                            function=bare,
                        )

                formulas[location] = formula

    return formulas, value_cells


def _check_circular(formulas: dict[str, str], sheet_names: list[str]) -> None:
    """Detect reference cycles among formula cells.

    Only formula cells can carry a cycle, because a constant has no outgoing
    edges. So ranges are resolved against the set of formula cells rather than
    expanded, which keeps this bounded by the number of formulas rather than
    by how large a range someone wrote.

    This is a deliberate small duplicate of the dependency graph in T06.
    Preflight runs before the graph is built, so it cannot use it.
    """
    if not formulas:
        return

    # Index formula cells by sheet so range containment is a cheap test.
    by_sheet: dict[str, list[tuple[int, int, str]]] = {name: [] for name in sheet_names}
    for address in formulas:
        sheet, coordinate = address.split("!", 1)
        min_col, min_row, _, _ = range_boundaries(coordinate)
        by_sheet.setdefault(sheet, []).append((min_row, min_col, address))

    graph = nx.DiGraph()
    graph.add_nodes_from(formulas)

    for address, formula in formulas.items():
        own_sheet = address.split("!", 1)[0]
        scannable = _strip_string_literals(formula)

        for match in _REFERENCE.finditer(scannable):
            reference = match.group(0)
            start, _, end = reference.partition(":")
            sheet, first = _split_reference(start)
            sheet = sheet or own_sheet
            if sheet not in by_sheet:
                continue

            last = _split_reference(end)[1] if end else first
            try:
                min_col, min_row, max_col, max_row = range_boundaries(f"{first}:{last}")
            except ValueError:
                continue

            for row, col, precedent in by_sheet[sheet]:
                if min_row <= row <= max_row and min_col <= col <= max_col:
                    graph.add_edge(precedent, address)

    try:
        cycle = nx.find_cycle(graph)
    except nx.NetworkXNoCycle:
        return

    path = " -> ".join(edge[0] for edge in cycle)
    raise PreflightRejected(
        Reason.CIRCULAR_REFERENCE,
        f"{path} -> {cycle[-1][1]} forms a loop. The recompute engine "
        "evaluates in dependency order, which a loop has no answer for.",
        location=cycle[0][0],
    )


# --- entry point -----------------------------------------------------------


def preflight(path: str | Path) -> PreflightReport:
    """Accept a workbook, or raise PreflightRejected naming the reason.

    Checks run cheapest first: the zip container, then a single pass over
    every cell, then cycle detection over the formulas that pass. The first
    problem found is the one reported.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(path)

    _check_container(path)

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        sheet_names = list(workbook.sheetnames)
        formulas, value_cells = _check_formulas(workbook)
    finally:
        workbook.close()

    _check_circular(formulas, sheet_names)

    return PreflightReport(
        path=path,
        sheet_names=sheet_names,
        formula_count=len(formulas),
        value_cell_count=value_cells,
    )
