"""The two tools the adjudicator has.

Signatures are in docs/ARCHITECTURE.md section 5:

    recompute_with_patch(cell, proposed_formula) -> {output: delta}
    inspect_range(sheet, range)                  -> cells with formulas

`recompute_with_patch` is the whole design in one function. The model states a
hypothesis, deterministic code tests it, and the result decides whether the
finding survives. The model does not get to assert an impact, it has to earn
one, and the reporter later drops any figure with no matching result here.

The tools return errors as values rather than raising. A model that proposes a
formula creating a circular reference should get told so and be able to try
again, not end the run.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from materia.formula import UnsupportedFormula
from materia.llm import ToolCall, ToolDefinition
from materia.parse import (
    InvalidReference,
    normalise_address,
    parse_reference,
    range_addresses,
)
from materia.recompute import CircularReference, EvaluationError, ExcelError, Model

# An inspection that returned a thousand cells would not be context, it would
# be the workbook again.
MAX_INSPECT_CELLS = 200


TOOL_DEFINITIONS = [
    ToolDefinition(
        name="recompute_with_patch",
        description=(
            "Apply a proposed formula to a copy of the model, recompute it, and "
            "return the change in each declared output cell. Use this to test a "
            "hypothesis. You may call it more than once. The model file is never "
            "modified."
        ),
        parameters={
            "type": "object",
            "properties": {
                "cell": {
                    "type": "string",
                    "description": "The cell to patch, as Sheet!Cell, for example 'P&L!AA15'.",
                },
                "proposed_formula": {
                    "type": "string",
                    "description": (
                        "The formula you believe should be there, starting with '='. "
                        "A plain number is also accepted."
                    ),
                },
            },
            "required": ["cell", "proposed_formula"],
        },
    ),
    ToolDefinition(
        name="inspect_range",
        description=(
            "Return the formulas, values, row labels and cell comments in a range. "
            "Use it when the peer group you were given is not enough context."
        ),
        parameters={
            "type": "object",
            "properties": {
                "sheet": {"type": "string", "description": "The sheet name."},
                "range": {
                    "type": "string",
                    "description": "An A1 range such as 'A14:H16', or a single cell.",
                },
            },
            "required": ["sheet", "range"],
        },
    ),
]


@dataclass(frozen=True)
class CellFacts:
    """Everything a reader can see about one cell."""

    address: str
    formula: str | None
    value: Any
    comment: str | None

    def as_dict(self) -> dict[str, Any]:
        record: dict[str, Any] = {"cell": self.address}
        if self.formula is not None:
            record["formula"] = self.formula
        if self.value is not None:
            record["value"] = self.value
        if self.comment:
            record["comment"] = self.comment
        return record


class Toolbox:
    """The tools bound to one workbook.

    The workbook is read once. Every patch is measured against the same
    baseline, so calling the tool repeatedly cannot drift.
    """

    def __init__(self, path: str | Path, outputs: list[str]) -> None:
        self.path = Path(path)
        self.outputs = [normalise_address(output) for output in outputs]
        self.model = Model.load(self.path, outputs=self.outputs)
        self.cells = _read_cells(self.path)

    @property
    def definitions(self) -> list[ToolDefinition]:
        return list(TOOL_DEFINITIONS)

    # --- the tools ---

    def recompute_with_patch(self, cell: str, proposed_formula: str) -> dict[str, Any]:
        """Test a hypothesis and return the true impact on each output."""
        sheet, separator, coordinate = str(cell).rpartition("!")
        if not separator:
            return {"error": f"{cell!r} is not a cell address like Sheet!A1"}
        try:
            parse_reference(coordinate)
        except InvalidReference:
            return {"error": f"{cell!r} is not a cell address like Sheet!A1"}

        address = normalise_address(str(cell))
        if address not in self.model.values:
            return {"error": f"{address} is not a populated cell in this workbook"}

        patch: str | float = proposed_formula
        if isinstance(proposed_formula, str) and not proposed_formula.startswith("="):
            try:
                patch = float(proposed_formula)
            except ValueError:
                patch = proposed_formula

        try:
            result = self.model.patch(address, patch)
        except UnsupportedFormula as error:
            return {"error": f"that formula is outside the supported grammar: {error}"}
        except CircularReference as error:
            return {"error": f"that formula creates a circular reference: {error}"}
        except EvaluationError as error:
            return {"error": str(error)}

        deltas: dict[str, Any] = {}
        for output, change in result.outputs.items():
            if change.delta is None:
                deltas[output] = _describe_broken(change)
            else:
                deltas[output] = round(change.delta, 6)
        return deltas

    def inspect_range(self, sheet: str, range: str) -> dict[str, Any]:  # noqa: A002
        """Return what is actually in a range, comments included."""
        if sheet not in self.cells:
            return {
                "error": f"no sheet named {sheet!r}",
                "sheets": sorted(self.cells),
            }

        text = str(range).replace("$", "").upper()
        start, _, end = text.partition(":")
        try:
            first = parse_reference(start)
            last = parse_reference(end) if end else first
        except InvalidReference:
            return {"error": f"{range!r} is not a range like A1:H16"}

        addresses = list(range_addresses(first, last, sheet))
        if len(addresses) > MAX_INSPECT_CELLS:
            return {
                "error": (
                    f"that range covers {len(addresses)} cells, above the "
                    f"{MAX_INSPECT_CELLS} cell limit. Ask for a smaller one."
                )
            }

        found = [
            self.cells[sheet][address].as_dict()
            for address in addresses
            if address in self.cells[sheet]
        ]
        return {"sheet": sheet, "range": text, "cells": found}

    # --- dispatch ---

    def run(self, call: ToolCall) -> dict[str, Any]:
        """Run one tool call from the model.

        A wrong tool name or a missing argument is answered rather than
        raised, because it is the model's mistake to correct.
        """
        handlers = {
            "recompute_with_patch": self.recompute_with_patch,
            "inspect_range": self.inspect_range,
        }
        handler = handlers.get(call.name)
        if handler is None:
            return {"error": f"no tool named {call.name!r}", "tools": sorted(handlers)}
        try:
            return handler(**call.arguments)
        except TypeError as error:
            return {"error": f"wrong arguments for {call.name}: {error}"}


def _describe_broken(change) -> str:
    """An output that stopped being a number has no delta to report."""
    after = change.after
    if isinstance(after, ExcelError):
        return f"the output becomes {after.value}"
    return f"the output stops being a number and becomes {after!r}"


def _read_cells(path: Path) -> dict[str, dict[str, CellFacts]]:
    """Read formulas, values and comments.

    Not in read only mode: openpyxl does not load comments there, and the
    comment on a deliberate override is exactly the evidence that separates it
    from a mistake.
    """
    formulas = openpyxl.load_workbook(path, data_only=False)
    values = openpyxl.load_workbook(path, data_only=True)
    try:
        sheets: dict[str, dict[str, CellFacts]] = {}
        for name in formulas.sheetnames:
            sheets[name] = {}
            for row in formulas[name].iter_rows():
                for cell in row:
                    if cell.value is None and cell.comment is None:
                        continue
                    address = f"{name}!{get_column_letter(cell.column)}{cell.row}"
                    is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
                    sheets[name][address] = CellFacts(
                        address=address,
                        formula=cell.value if is_formula else None,
                        value=values[name][cell.coordinate].value,
                        comment=cell.comment.text if cell.comment else None,
                    )
        return sheets
    finally:
        formulas.close()
        values.close()
