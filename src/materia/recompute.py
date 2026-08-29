"""Deterministic recompute engine over the supported grammar.

This is the load bearing component of the whole project. Every impact figure
in a report and the ground truth materiality of every seeded mutation both
come from here, so the semantics below follow Excel rather than Python
wherever the two disagree, and each disagreement is called out where it
happens.

The engine evaluates a workbook in topological order, applies a single cell
patch, recomputes, and returns the change on each declared output. It never
writes to the workbook it was loaded from.

Excel behaviour implemented deliberately, because the obvious Python version
is wrong in each case:

  ROUND rounds half away from zero. Python rounds half to even, so
      round(2.5) is 2 while Excel's ROUND(2.5,0) is 3.
  SUM, AVERAGE, MIN and MAX skip text and booleans found inside a range, but
      coerce them when passed directly as arguments.
  Empty cells are 0 in arithmetic and are skipped by the aggregates.
  Text that looks like a number is coerced in arithmetic, so "5"+1 is 6.
  Division by zero produces a value, not an exception, and propagates.
"""

from __future__ import annotations

from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Iterator

import networkx as nx
import openpyxl
from openpyxl.utils import get_column_letter

from materia.formula import (
    BinaryOp,
    Boolean,
    CellRef,
    FunctionCall,
    Node,
    Number,
    Percent,
    RangeRef,
    Text,
    UnaryOp,
    parse_formula,
)
from materia.parse import Reference

# A range wider than this is refused rather than enumerated. Real models do
# not aggregate over a million cells, and silently taking minutes would be
# worse than saying no.
MAX_RANGE_CELLS = 65_536


class ExcelError(str, Enum):
    """Excel's error values, which are values rather than exceptions.

    They propagate through arithmetic exactly as they do in a spreadsheet, so
    a bad patch produces #DIV/0! in an output rather than crashing a run.
    """

    DIV0 = "#DIV/0!"
    VALUE = "#VALUE!"
    NUM = "#NUM!"
    REF = "#REF!"


Scalar = float | str | bool | None | ExcelError


class RangeValue(tuple):
    """A rectangle of cell values, in row then column order.

    Distinct from a plain tuple so a range used where a single value belongs,
    such as `=A1:A3*2`, can be refused rather than silently reduced.
    """


class CircularReference(Exception):
    """A patch produced a workbook that refers to itself."""


class EvaluationError(Exception):
    """The engine cannot evaluate this workbook at all."""


# --- value coercion --------------------------------------------------------


def _as_number(value: Scalar) -> float | ExcelError:
    """Coerce to a number the way Excel does in an arithmetic context."""
    if isinstance(value, ExcelError):
        return value
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)  # Excel coerces numeric text: "5"+1 is 6
    except ValueError:
        return ExcelError.VALUE


def _as_bool(value: Scalar) -> bool | ExcelError:
    if isinstance(value, ExcelError):
        return value
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    return ExcelError.VALUE


def _type_rank(value: Scalar) -> int:
    """Excel orders across types: number < text < FALSE < TRUE."""
    if isinstance(value, bool):
        return 2
    if isinstance(value, (int, float)) or value is None:
        return 0
    return 1


def _compare(left: Scalar, right: Scalar) -> int | ExcelError:
    """Three way comparison following Excel's rules.

    An empty cell equals 0 and equals the empty string, text comparison is
    case insensitive, and values of different types order by type.
    """
    if isinstance(left, ExcelError):
        return left
    if isinstance(right, ExcelError):
        return right

    if left is None:
        left = "" if isinstance(right, str) else 0.0
    if right is None:
        right = "" if isinstance(left, str) else 0.0

    left_rank, right_rank = _type_rank(left), _type_rank(right)
    if left_rank != right_rank:
        return -1 if left_rank < right_rank else 1

    if isinstance(left, bool):
        return (left > right) - (left < right)
    if isinstance(left, str):
        lowered, other = left.lower(), str(right).lower()
        return (lowered > other) - (lowered < other)
    return (float(left) > float(right)) - (float(left) < float(right))


def _excel_round(value: float, digits: int) -> float:
    """Round half away from zero, which is what Excel does.

    Python's round() rounds half to even, so round(2.5) is 2 where Excel's
    ROUND(2.5,0) is 3. Getting this wrong would bias every rounded figure in
    the corpus in a way nobody would notice.
    """
    factor = 10.0**digits
    scaled = value * factor

    # The nudge absorbs binary representation error: 2.675*100 comes out as
    # 267.49999999999997 and would otherwise round down to 2.67, where Excel
    # gives 2.68. It is proportional to magnitude rather than a fixed amount,
    # because a fixed epsilon large enough to fix 2.675 also pushes a value
    # genuinely just below the boundary, such as 0.4999999995, up to 1.
    nudge = abs(scaled) * 1e-12
    nudged = scaled + (nudge if scaled >= 0 else -nudge)

    rounded = float(int(nudged + (0.5 if nudged >= 0 else -0.5)))
    return rounded / factor


# --- the model -------------------------------------------------------------


@dataclass(frozen=True)
class OutputDelta:
    """What a patch did to one declared output."""

    address: str
    before: Scalar
    after: Scalar
    delta: float | None
    relative: float | None


@dataclass(frozen=True)
class PatchResult:
    address: str
    patch: str
    outputs: dict[str, OutputDelta]

    def as_tool_result(self) -> dict[str, float | None]:
        """The `{output: delta}` shape the agent tool returns."""
        return {address: output.delta for address, output in self.outputs.items()}


class Model:
    """A workbook loaded as constants plus parsed formulas.

    Run preflight first. This assumes every formula is inside the supported
    grammar and that the workbook has no cycles.
    """

    def __init__(
        self,
        constants: dict[str, Scalar],
        formulas: dict[str, Node],
        outputs: list[str] | None = None,
    ) -> None:
        self.constants = dict(constants)
        self.formulas = dict(formulas)
        self.outputs = [_normalise_address(address) for address in (outputs or [])]
        self.values = self._evaluate_all(self.formulas)

    # --- construction ---

    @classmethod
    def load(cls, path: str | Path, outputs: list[str] | None = None) -> "Model":
        """Read a workbook read only. The file is never written."""
        workbook = openpyxl.load_workbook(Path(path), read_only=True, data_only=False)
        try:
            constants: dict[str, Scalar] = {}
            formulas: dict[str, Node] = {}
            for sheet_name in workbook.sheetnames:
                for row in workbook[sheet_name].iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        address = f"{sheet_name}!{cell.coordinate}"
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            formulas[address] = parse_formula(cell.value)
                        else:
                            constants[address] = cell.value
        finally:
            workbook.close()
        return cls(constants, formulas, outputs)

    @classmethod
    def from_cells(
        cls, cells: dict[str, str | float], outputs: list[str] | None = None
    ) -> "Model":
        """Build a model from a mapping of address to formula or value.

        For tests and for the corpus generator. A string starting with `=` is
        a formula, anything else is a constant.
        """
        constants: dict[str, Scalar] = {}
        formulas: dict[str, Node] = {}
        for address, content in cells.items():
            key = _normalise_address(address)
            if isinstance(content, str) and content.startswith("="):
                formulas[key] = parse_formula(content)
            else:
                constants[key] = content
        return cls(constants, formulas, outputs)

    # --- evaluation ---

    def _evaluate_all(self, formulas: dict[str, Node]) -> dict[str, Scalar]:
        values: dict[str, Scalar] = dict(self.constants)
        for address in self._topological_order(formulas):
            sheet = address.split("!", 1)[0]
            result = _Evaluator(values, formulas.keys() | self.constants.keys(), sheet).run(
                formulas[address]
            )
            values[address] = result if not isinstance(result, RangeValue) else ExcelError.VALUE
        return values

    def _topological_order(self, formulas: dict[str, Node]) -> list[str]:
        known = formulas.keys() | self.constants.keys()
        graph = nx.DiGraph()
        graph.add_nodes_from(formulas)
        for address, node in formulas.items():
            sheet = address.split("!", 1)[0]
            for precedent in _precedents(node, sheet, known):
                if precedent in formulas:
                    graph.add_edge(precedent, address)
        try:
            return list(nx.topological_sort(graph))
        except nx.NetworkXUnfeasible as unfeasible:
            cycle = nx.find_cycle(graph)
            path = " -> ".join(edge[0] for edge in cycle)
            raise CircularReference(f"{path} -> {cycle[-1][1]}") from unfeasible

    # --- reading ---

    def value(self, address: str) -> Scalar:
        return self.values.get(_normalise_address(address))

    # --- patching ---

    def patch(self, address: str, replacement: str | float) -> PatchResult:
        """Apply one cell change to a copy and report the effect on outputs.

        `replacement` is either a formula starting with `=` or a literal
        value. The model itself is not modified, so a Model can be patched
        repeatedly and every result is measured against the same baseline.
        """
        if not self.outputs:
            raise EvaluationError("no declared output cells, so there is nothing to measure")

        address = _normalise_address(address)
        formulas = dict(self.formulas)
        constants_backup = self.constants

        is_formula = isinstance(replacement, str) and replacement.startswith("=")
        try:
            if is_formula:
                formulas[address] = parse_formula(replacement)
                self.constants = {
                    key: value for key, value in self.constants.items() if key != address
                }
            else:
                formulas.pop(address, None)
                self.constants = {**self.constants, address: replacement}
            patched = self._evaluate_all(formulas)
        finally:
            self.constants = constants_backup

        outputs = {}
        for output in self.outputs:
            before = self.values.get(output)
            after = patched.get(output)
            outputs[output] = OutputDelta(
                address=output,
                before=before,
                after=after,
                delta=_delta(before, after),
                relative=_relative(before, after),
            )
        return PatchResult(address=address, patch=str(replacement), outputs=outputs)


def _delta(before: Scalar, after: Scalar) -> float | None:
    """None rather than 0.0 when either side is not a number.

    An output that turned into #DIV/0! has not moved by zero, and reporting it
    as unchanged would be a lie the whole design exists to prevent.
    """
    if isinstance(before, (bool, ExcelError)) or isinstance(after, (bool, ExcelError)):
        return None
    if not isinstance(before, (int, float)) or not isinstance(after, (int, float)):
        return None
    return float(after) - float(before)


def _relative(before: Scalar, after: Scalar) -> float | None:
    delta = _delta(before, after)
    if delta is None or not isinstance(before, (int, float)) or float(before) == 0.0:
        return None
    return delta / abs(float(before))


# --- reference resolution --------------------------------------------------


def _normalise_address(address: str) -> str:
    sheet, _, coordinate = address.rpartition("!")
    return f"{sheet.strip(chr(39))}!{coordinate.replace('$', '').upper()}"


def _cell_address(reference: Reference, default_sheet: str) -> str:
    sheet = (reference.sheet or default_sheet).strip("'")
    return f"{sheet}!{get_column_letter(reference.column)}{reference.row}"


def _range_addresses(node: RangeRef, default_sheet: str) -> Iterator[str]:
    """Every address in the rectangle, row then column order.

    Empty cells are included so SUMIF can line a criteria range up with a sum
    range of the same shape.
    """
    sheet = (node.start.sheet or default_sheet).strip("'")
    first_row, last_row = sorted((node.start.row, node.end.row))
    first_column, last_column = sorted((node.start.column, node.end.column))

    size = (last_row - first_row + 1) * (last_column - first_column + 1)
    if size > MAX_RANGE_CELLS:
        raise EvaluationError(
            f"range covers {size} cells, above the {MAX_RANGE_CELLS} limit"
        )

    for row in range(first_row, last_row + 1):
        for column in range(first_column, last_column + 1):
            yield f"{sheet}!{get_column_letter(column)}{row}"


def _precedents(node: Node, sheet: str, known: set[str]) -> Iterator[str]:
    """Addresses this formula reads, restricted to cells that exist."""
    from materia.formula import references

    for reference in references(node):
        if isinstance(reference, CellRef):
            address = _cell_address(reference.reference, sheet)
            if address in known:
                yield address
        else:
            for address in _range_addresses(reference, sheet):
                if address in known:
                    yield address


# --- evaluator -------------------------------------------------------------


class _Evaluator:
    def __init__(self, values: dict[str, Scalar], known: set[str], sheet: str) -> None:
        self._values = values
        self._known = known
        self._sheet = sheet

    def run(self, node: Node) -> Scalar | RangeValue:
        return self._evaluate(node)

    def _evaluate(self, node: Node) -> Scalar | RangeValue:
        if isinstance(node, Number):
            return node.value
        if isinstance(node, Text):
            return node.value
        if isinstance(node, Boolean):
            return node.value
        if isinstance(node, CellRef):
            return self._values.get(_cell_address(node.reference, self._sheet))
        if isinstance(node, RangeRef):
            return RangeValue(
                self._values.get(address)
                for address in _range_addresses(node, self._sheet)
            )
        if isinstance(node, UnaryOp):
            return self._unary(node)
        if isinstance(node, Percent):
            value = _as_number(self._scalar(self._evaluate(node.operand)))
            return value if isinstance(value, ExcelError) else value / 100.0
        if isinstance(node, BinaryOp):
            return self._binary(node)
        if isinstance(node, FunctionCall):
            return self._call(node)
        raise EvaluationError(f"cannot evaluate {node!r}")

    def _scalar(self, value: Scalar | RangeValue) -> Scalar:
        """A range used where one value belongs is an error, not a guess."""
        return ExcelError.VALUE if isinstance(value, RangeValue) else value

    def _unary(self, node: UnaryOp) -> Scalar:
        value = _as_number(self._scalar(self._evaluate(node.operand)))
        if isinstance(value, ExcelError):
            return value
        return -value if node.operator == "-" else value

    def _binary(self, node: BinaryOp) -> Scalar:
        left = self._scalar(self._evaluate(node.left))
        right = self._scalar(self._evaluate(node.right))

        if node.operator in ("=", "<>", "<", "<=", ">", ">="):
            comparison = _compare(left, right)
            if isinstance(comparison, ExcelError):
                return comparison
            return {
                "=": comparison == 0,
                "<>": comparison != 0,
                "<": comparison < 0,
                "<=": comparison <= 0,
                ">": comparison > 0,
                ">=": comparison >= 0,
            }[node.operator]

        first, second = _as_number(left), _as_number(right)
        if isinstance(first, ExcelError):
            return first
        if isinstance(second, ExcelError):
            return second

        if node.operator == "+":
            return first + second
        if node.operator == "-":
            return first - second
        if node.operator == "*":
            return first * second
        if node.operator == "/":
            return ExcelError.DIV0 if second == 0 else first / second
        if node.operator == "^":
            if first < 0 and second != int(second):
                return ExcelError.NUM
            if first == 0 and second < 0:
                return ExcelError.DIV0
            try:
                return float(first**second)
            except (OverflowError, ValueError):
                return ExcelError.NUM
        raise EvaluationError(f"unknown operator {node.operator!r}")

    # --- functions ---

    def _call(self, node: FunctionCall) -> Scalar:
        handler = getattr(self, f"_function_{node.name.lower()}")
        return handler([self._evaluate(argument) for argument in node.arguments])

    def _numbers(self, arguments: list[Scalar | RangeValue]) -> list[float] | ExcelError:
        """Collect the numbers an aggregate should see.

        Inside a range, text and booleans are skipped and empties ignored.
        Passed directly as an argument they are coerced instead. That
        asymmetry is Excel's, and SUM(A1:A3) versus SUM(TRUE) depends on it.
        """
        collected: list[float] = []
        for argument in arguments:
            if isinstance(argument, RangeValue):
                for value in argument:
                    if isinstance(value, ExcelError):
                        return value
                    if isinstance(value, bool) or value is None:
                        continue
                    if isinstance(value, (int, float)):
                        collected.append(float(value))
                continue
            number = _as_number(argument)
            if isinstance(number, ExcelError):
                return number
            collected.append(number)
        return collected

    def _function_sum(self, arguments):
        numbers = self._numbers(arguments)
        return numbers if isinstance(numbers, ExcelError) else sum(numbers)

    def _function_average(self, arguments):
        numbers = self._numbers(arguments)
        if isinstance(numbers, ExcelError):
            return numbers
        return ExcelError.DIV0 if not numbers else sum(numbers) / len(numbers)

    def _function_min(self, arguments):
        numbers = self._numbers(arguments)
        if isinstance(numbers, ExcelError):
            return numbers
        return min(numbers) if numbers else 0.0  # Excel gives 0 for an empty range

    def _function_max(self, arguments):
        numbers = self._numbers(arguments)
        if isinstance(numbers, ExcelError):
            return numbers
        return max(numbers) if numbers else 0.0

    def _function_abs(self, arguments):
        value = _as_number(self._scalar(arguments[0]))
        return value if isinstance(value, ExcelError) else abs(value)

    def _function_round(self, arguments):
        value = _as_number(self._scalar(arguments[0]))
        digits = _as_number(self._scalar(arguments[1]))
        if isinstance(value, ExcelError):
            return value
        if isinstance(digits, ExcelError):
            return digits
        return _excel_round(value, int(digits))

    def _function_if(self, arguments):
        condition = _as_bool(self._scalar(arguments[0]))
        if isinstance(condition, ExcelError):
            return condition
        if condition:
            return self._scalar(arguments[1])
        # Excel returns FALSE when the third argument is left out.
        return self._scalar(arguments[2]) if len(arguments) > 2 else False

    def _function_sumif(self, arguments):
        criteria_range = arguments[0]
        criterion = self._scalar(arguments[1])
        sum_range = arguments[2] if len(arguments) > 2 else criteria_range

        if not isinstance(criteria_range, RangeValue):
            criteria_range = RangeValue([self._scalar(criteria_range)])
        if not isinstance(sum_range, RangeValue):
            sum_range = RangeValue([self._scalar(sum_range)])

        matches = _criteria_test(criterion)
        total = 0.0
        for index, value in enumerate(criteria_range):
            if index >= len(sum_range):
                break  # Excel takes the shape of the criteria range
            if not matches(value):
                continue
            target = sum_range[index]
            if isinstance(target, ExcelError):
                return target
            if isinstance(target, bool) or target is None:
                continue
            if isinstance(target, (int, float)):
                total += float(target)
        return total


def _criteria_test(criterion: Scalar):
    """Build the match test for a SUMIF criterion.

    A criterion is either a bare value meaning equality, or a string carrying
    a leading comparison operator such as ">120" or "<>0".
    """
    operator = "="
    target: Scalar = criterion

    if isinstance(criterion, str):
        for candidate in ("<>", ">=", "<=", ">", "<", "="):
            if criterion.startswith(candidate):
                operator = candidate
                remainder = criterion[len(candidate) :]
                try:
                    target = float(remainder)
                except ValueError:
                    target = remainder
                break

    def matches(value: Scalar) -> bool:
        comparison = _compare(value, target)
        if isinstance(comparison, ExcelError):
            return False
        return {
            "=": comparison == 0,
            "<>": comparison != 0,
            "<": comparison < 0,
            "<=": comparison <= 0,
            ">": comparison > 0,
            ">=": comparison >= 0,
        }[operator]

    return matches
