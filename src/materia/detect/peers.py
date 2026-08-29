"""Peer groups: the cells a formula should look like.

A peer group is a run of cells along one axis that were filled from the same
origin, so in R1C1 they should all carry an identical token. The whole
detection layer rests on that: a cell that does not match its peers is the
only structural signal a spreadsheet gives you. See docs/ARCHITECTURE.md
section 2.
"""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from pathlib import Path

from materia.parse import FormulaCell, read_formulas

# Below this many cells a "mode" means nothing: three cells where one differs
# is not a pattern with an exception, it is three cells.
MINIMUM_PEERS = 4


@dataclass(frozen=True)
class PeerCell:
    """One member of a peer group, as evidence in a candidate."""

    address: str
    formula: str
    r1c1: str


@dataclass(frozen=True)
class PeerGroup:
    """A row or column of cells that should share one R1C1 token."""

    sheet: str
    axis: str  # "row" or "column"
    index: int
    members: tuple[PeerCell, ...]

    @property
    def tokens(self) -> Counter:
        return Counter(member.r1c1 for member in self.members)

    @property
    def mode(self) -> str:
        """The token most of the group agrees on."""
        return self.tokens.most_common(1)[0][0]

    @property
    def mode_share(self) -> float:
        return self.tokens[self.mode] / len(self.members)

    def conforming(self, limit: int = 3) -> tuple[PeerCell, ...]:
        """A few members that do match the mode, to show as evidence."""
        return tuple(m for m in self.members if m.r1c1 == self.mode)[:limit]

    def describe(self) -> str:
        return f"{self.sheet} {self.axis} {self.index}"


@dataclass
class Sheet:
    """Every populated cell on one sheet, formula or not."""

    name: str
    formulas: dict[tuple[int, int], FormulaCell]
    values: set[tuple[int, int]]

    def occupied(self, row: int, column: int) -> bool:
        return (row, column) in self.formulas or (row, column) in self.values


@dataclass
class Workbook:
    """What the detectors read: formulas, values and peer groups."""

    path: Path
    sheets: dict[str, Sheet]
    groups: tuple[PeerGroup, ...]


def load(path: str | Path) -> Workbook:
    """Read a workbook into the shape the detectors work on."""
    import openpyxl

    path = Path(path)
    sheets: dict[str, Sheet] = {}

    book = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        for name in book.sheetnames:
            sheets[name] = Sheet(name=name, formulas={}, values=set())
            for row in book[name].iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        continue
                    sheets[name].values.add((cell.row, cell.column))
    finally:
        book.close()

    for cell in read_formulas(path):
        sheets[cell.sheet].formulas[(cell.row, cell.column)] = cell

    return Workbook(path=path, sheets=sheets, groups=tuple(_build_groups(sheets)))


def _build_groups(sheets: dict[str, Sheet]) -> list[PeerGroup]:
    """Every row and column run of formulas long enough to have a pattern."""
    groups = []
    for sheet in sheets.values():
        by_row: dict[int, list[FormulaCell]] = {}
        by_column: dict[int, list[FormulaCell]] = {}
        for (row, column), cell in sheet.formulas.items():
            by_row.setdefault(row, []).append(cell)
            by_column.setdefault(column, []).append(cell)

        for axis, buckets, key in (
            ("row", by_row, lambda c: c.column),
            ("column", by_column, lambda c: c.row),
        ):
            for index, cells in buckets.items():
                if len(cells) < MINIMUM_PEERS:
                    continue
                members = tuple(
                    PeerCell(cell.address, cell.formula, cell.r1c1)
                    for cell in sorted(cells, key=key)
                )
                groups.append(PeerGroup(sheet.name, axis, index, members))
    return groups
